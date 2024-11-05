
from rest_framework.pagination import PageNumberPagination
from django_filters import rest_framework as filters
from django.db.models import Prefetch, OuterRef, Subquery
from django.contrib.auth import logout
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from rest_framework import status,generics, viewsets,response
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.urls import reverse
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.db.models import Count, Case, When, IntegerField,Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.generics import UpdateAPIView, RetrieveUpdateAPIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django.db import connection,DatabaseError
from rest_framework import serializers
from rest_framework.views import APIView
from rest_framework import generics
from .serializer import *
from userprofile.models import *
from .models import User
from django.contrib.auth import get_user_model
from rest_framework.exceptions import NotFound,ValidationError
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
import pandas as pd
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse,HttpResponse,Http404
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
import re
from dateutil import parser
import numpy as np
import os
from rest_framework.parsers import MultiPartParser
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import logging

# Configure logging
logging.basicConfig(filename='debug.log', level=logging.DEBUG)

User = get_user_model()


class CustomPagination(PageNumberPagination):
    page_size = 600
    page_size_query_param = 'page_size'
    max_page_size = 1000


# Create your views here.

# User data

@api_view(['GET']) 
@permission_classes([IsAuthenticated])
def get_users(request):
        try:
            if request.query_params:
                user=User.objects.filter(**request.query_params.dict())
            else:
                user = User.objects.all()
        
            # if there is something in items else raise error
            if user:
                serializer = UserSerializer(user, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
class AluminiBulkRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]
    def post(self, request):
        serializer = AlumniBulkRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AluminiRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]
    def post(self, request):
        print(request.data)
        serializer = AlumniRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            if request.query_params:
                user = User.objects.filter(**request.query_params.dict(), is_alumni=True)
            else:
                user = User.objects.filter(is_alumni=True)

            # if there is something in items else raise error
            if user:
                serializer = AlumniSerializer(user, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
class AluminiListView(APIView):
    permission_classes = [IsAuthenticated, ]
    
    def get(self, request):
        try:
            query_params = request.query_params.dict()
            query_param_key = list(query_params.keys())[0] if query_params else None
            query_param_value = query_params.get(query_param_key)

            base_query = '''WITH distinct_careers AS (
                                SELECT DISTINCT ON (userprofile_employment.career, userprofile_alumni.id)
                                    userprofile_alumni.id as alumn_id, userprofile_employment.career, userprofile_employment.on_going
                                FROM userprofile_employment
                                JOIN userprofile_alumni ON userprofile_employment.alumn_id = userprofile_alumni.id
                                ORDER BY userprofile_alumni.id, userprofile_employment.career, userprofile_employment.on_going DESC
                            )
                            SELECT api_user.id, api_user.email, api_user.image_url, api_user.first_name, api_user.last_name, api_user.phone1, 
                                userprofile_alumni.reg_number, userprofile_alumni.id as alumn_id, userprofile_grade.grade_name, 
                                userprofile_grade.id AS grade_id, userprofile_family.family_name, userprofile_family.id AS family_id, 
                                userprofile_combination.combination_name, userprofile_combination.id AS combination_id, 
                                STRING_AGG(distinct_careers.career, ', ' ORDER BY distinct_careers.on_going DESC) AS career
                            FROM api_user
                            LEFT OUTER JOIN userprofile_alumni ON api_user.id = userprofile_alumni.user_id
                            LEFT OUTER JOIN userprofile_family ON userprofile_alumni.family_id = userprofile_family.id
                            LEFT OUTER JOIN userprofile_grade ON userprofile_family.grade_id = userprofile_grade.id
                            LEFT OUTER JOIN userprofile_combination ON userprofile_alumni.combination_id = userprofile_combination.id
                            LEFT OUTER JOIN distinct_careers ON userprofile_alumni.id = distinct_careers.alumn_id
                            WHERE api_user.is_alumni'''

            group_by_clause = ''' GROUP BY api_user.id, api_user.email, api_user.image_url, api_user.first_name, api_user.last_name, api_user.phone1, 
                                        userprofile_alumni.reg_number, userprofile_alumni.id, userprofile_grade.grade_name, userprofile_grade.id, 
                                        userprofile_family.family_name, userprofile_family.id, userprofile_combination.combination_name, userprofile_combination.id'''

            if query_param_key == "grade_id":
                query = f"{base_query} AND userprofile_grade.id = %s{group_by_clause}"
            elif query_param_key == "family_id":
                query = f"{base_query} AND userprofile_family.id = %s{group_by_clause}"
            elif query_param_key == "combination_id":
                query = f"{base_query} AND userprofile_combination.id = %s{group_by_clause}"
            else:
                query = f"{base_query}{group_by_clause}"
                query_param_value = None

            with connection.cursor() as cursor:
                cursor.execute(query, [query_param_value] if query_param_value else [])
                data1 = cursor.fetchall()
                
                data = []
            if data1:
                for i in data1:
                    data.append({
                        'id':i[0],
                        'email':i[1],
                        'image_url':i[2],
                        'first_name': i[3],
                        'last_name': i[4],
                        'phone1': i[5],
                        'reg_number': i[6],
                        'alumn_id': i[7],
                        'grade_name': i[8],
                        'grade_id': i[9],
                        'family_name': i[10],
                        'family_id':i[11],
                        'combination_name': i[12],
                        'combination_id': i[13],
                        'career': i[14]
                    })

            serializer = AlumniListsSerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except KeyError:
            return Response({'error': 'Invalid query parameter.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
class AluminiListByEyView(APIView):
    permission_classes = [IsAuthenticated, ]
    
    def get(self,request):
        try:
            if request.query_params:
                if(list(request.query_params.keys())[0]=="ep_id"):
                    user = User.objects.raw('''select api_user.id,api_user.email, api_user.image_url, api_user.first_name, api_user.last_name, api_user.phone1, userprofile_grade.grade_name,userprofile_grade.id as grade_id,userprofile_family.family_name, userprofile_family.id as family_id,userprofile_combination.combination_name,userprofile_combination.id as combination_id,userprofile_ep.id as ep_id,userprofile_ep.title as ep_title from api_user left outer join userprofile_alumni on api_user.id=userprofile_alumni.user_id left outer join userprofile_family on userprofile_alumni.family_id=userprofile_family.id left outer join userprofile_grade on userprofile_family.grade_id=userprofile_grade.id left outer join userprofile_combination on userprofile_alumni.combination_id=userprofile_combination.id left outer join userprofile_alumni_eps on userprofile_alumni.id=userprofile_alumni_eps.alumni_id left outer join userprofile_ep on userprofile_ep.id=userprofile_alumni_eps.ep_id where api_user.is_alumni and userprofile_ep.id=%s;''',[request.query_params.dict()["ep_id"]])
                    serializer = AlumniListbyEPSerializer(user, many=True)
                    return Response(serializer.data)
                else:
                    return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        

class StaffUserView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = StaffUserRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            if request.query_params:
                query_params = request.query_params.dict()
                crc = User.objects.filter(
                    Q(is_crc=True) |
                    Q(is_teacher=True) |
                    Q(is_librarian=True) |
                    Q(is_superuser=True),
                    **query_params
                )
            else:
                crc = User.objects.filter(
                    Q(is_crc=True) |
                    Q(is_teacher=True) |
                    Q(is_librarian=True) |
                    Q(is_superuser=True)
                )
            # if there is something in items else raise error
            if crc:
                serializer = StaffUserSerializer(crc, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_user(request, pk):
    user = User.objects.get(pk=pk)
    data = UpdateUserSerializer(instance=user, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        print(data.errors)
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_user_image(request, pk):
    user = User.objects.get(pk=pk)
    data = UpdateUserImageUrlSerializer(instance=user, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        print(data.errors)
        return Response(error=data.errors,status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_user_position(request, pk):
    user = CrcProfile.objects.get(user_id=pk)
    data = StaffRoleSerializer(instance=user, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        print(data.errors)
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_alumni_info(request):
    serializer = AlumniInfoRegSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    alumni_info = serializer.save()
    print(alumni_info.id)
    for ep_id in request.data.get('eps'):
        try:
            ep = Ep.objects.get(id=ep_id)
            alumni_info.eps.add(ep)
        except Ep.DoesNotExist:
            raise NotFound()

    return Response(data={"id":alumni_info.id},status=status.HTTP_201_CREATED)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_alumni_info(request, pk=None):
    alumn = Alumni.objects.get(pk=pk)

    serializer = AlumniInfoUpdateSerializer(alumn, data=request.data)
    if serializer.is_valid():
        serializer.save()

        eps = []

        for ep_id in request.data.get('eps'):
            try:
                ep = Ep.objects.get(id=ep_id)
                eps.append(ep)
            except Ep.DoesNotExist:
                raise NotFound()

        alumn.eps.set(eps)

        return Response(data=serializer.data, status=status.HTTP_200_OK)
    else:
        print(serializer.errors)
        return Response(data=serializer.errors, status=status.HTTP_200_OK)


#login logout and change and reset password portal
class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)

        # Add custom claims
        token['first_name'] = user.first_name
        token['last_name'] = user.last_name
        token['email'] = user.email
        token['phone1'] = user.phone1
        token['is_superuser'] = user.is_superuser
        token['is_crc'] = user.is_crc
        token['is_alumni'] = user.is_alumni
        token['is_librarian'] = user.is_librarian
        token['is_teacher'] = user.is_teacher
        token['is_student'] = user.is_student
        token['id'] = user.id
        # ...

        return token


class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


class LogoutView(APIView):
    def post(self, request):
        logout(request)
        return Response({'msg': 'Successfully Logged out'}, status=status.HTTP_200_OK)



class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = PasswordChangeSerializer(context={'request': request}, data=request.data)
        serializer.is_valid(raise_exception=True)
        request.user.set_password(serializer.validated_data['new_password'])
        request.user.save()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
class PasswordReset(generics.GenericAPIView):
    permission_classes = [IsAuthenticated, ]
    serializer_class = EmailSerilizer

    def post(self, request):

        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.data["email"]
        user = User.objects.filter(email=email).first()
        if user:
            encoded_pk = urlsafe_base64_encode(force_bytes(user.pk))
            token = PasswordResetTokenGenerator().make_token(user)

            reset_url = reverse(
                "reset-password",
                kwargs={"encoded_pk":encoded_pk, "token":token}
            )


            return Response(
                {
                    "message":reset_url
                },
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message":"User doesn't exists"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
class ResetPassword(generics.GenericAPIView):
    permission_classes = [IsAuthenticated, ]
    serializer_class = ResetPasswordSerializer

    def patch(self, request, *args, **kwargs):

        serializer = self.serializer_class(
            data=request.data, context={"kwargs":kwargs}
        )
        serializer.is_valid(raise_exception=True)

        return Response(
            {"message":"Password reset complete"},
            status=status.HTTP_200_OK,
        )
    
    # End login logout and change password portal


@api_view(['GET'])
def getRoutes(request):
    routes = [
        '/api/token/',
        '/api/register/',
        '/api/token/refresh/'
    ]
    return Response(routes)


# families and grades views

class GradeView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = GradeSerializers(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            if request.query_params:
                grades = Grade.objects.filter(**request.query_params.dict()).order_by('start_academic_year')
            else:
                grades = Grade.objects.all()

            # if there is something in items else raise error
            if grades:
                serializer = GradeSerializers(grades, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_grade(request, pk):
    grade = Grade.objects.get(pk=pk)
    data = GradeSerializers(instance=grade, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_family(request, pk):
    family = Family.objects.get(pk=pk)
    data = FamilySerializer(instance=family, data=request.data)
    if data.is_valid():
        data.save()
        return Response(data.data, status=status.HTTP_201_CREATED)
    else:
        return Response(data.errors, status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_grade(request, pk):
    grade = get_object_or_404(Grade, pk=pk)
    grade.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_family(request, pk):
    family = get_object_or_404(Family, pk=pk)
    family.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_families_to_grade(request):
    data = AddFamilySerializer(data=request.data)
    if data.is_valid():
        data.save()
        return Response(data.data, status=status.HTTP_201_CREATED)
    else:
        print(data.errors)
        return Response(data.errors, status=status.HTTP_404_NOT_FOUND)
    
class GradesAndFamiliesView(APIView):
    permission_classes = [IsAuthenticated, ]  
    def get(self,request):
        try:
            user = User.objects.raw("select userprofile_family.id,userprofile_family.family_name,userprofile_grade.grade_name,userprofile_grade.start_academic_year, userprofile_grade.end_academic_year from userprofile_grade inner join userprofile_family on userprofile_grade.id=userprofile_family.grade_id order by userprofile_grade.start_academic_year;")
    
            # if there is something in items else raise error
            if user:
                serializer = GradesAndFamiliesSerializer(user, many=True)
                return Response(serializer.data)
            else:
                return Response([])
            
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


# End


# Ep data means art,sport, sciences and clubs CRUD

class EpView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = EpSerializer(data=request.data)
        # validating for already existing data
        if Ep.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                ep = Ep.objects.filter(**request.query_params.dict())
            else:
                ep = Ep.objects.all()

            # if there is something in items else raise error
            if ep:
                serializer = EpSerializer(ep, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_Ep(request, pk):
    ep = Ep.objects.get(pk=pk)
    data = EpSerializer(instance=ep, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_ep(request, pk):
    ep = get_object_or_404(Ep, pk=pk)
    ep.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


# Combination data view

class CombinationRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = CombinationSerializer(data=request.data)
        # validating for already existing data
        if Combination.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                comb = Combination.objects.filter(**request.query_params.dict())
            else:
                comb = Combination.objects.all()

            # if there is something in items else raise error
            if comb:
                serializer = CombinationSerializer(comb, many=True)
                return Response(serializer.data)
            else:
                return Response([])
            
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_Comb(request, pk):
    comb = Combination.objects.get(pk=pk)
    data = CombinationSerializer(instance=comb, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_comb(request, pk):
    comb = get_object_or_404(Combination, pk=pk)
    comb.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


# end

class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all()
    serializer_class = EventSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAuthenticated]
        return super().get_permissions()

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Event.DoesNotExist:
            return Response({'error': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except Event.DoesNotExist:
            return Response({'error': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Event.DoesNotExist:
            return Response({'error': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# end


# Story data view
class StoryView(APIView):
    #permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = StorySerializer(data=request.data)
        # validating for already existing data
        if Story.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                onestory = Story.objects.filter(**request.query_params.dict())
                serializer = StoryWithAlumnSerializer(onestory, many=True)
                return Response(serializer.data)
            else:
                story = User.objects.raw(
                    "SELECT api_user.id as id, api_user.email as email, api_user.phone1 as phone1, api_user.first_name as first_name, api_user.last_name as last_name,api_user.image_url,userprofile_Story.description as description,userprofile_Story.image,userprofile_Story.video,userprofile_Story.title, userprofile_Story.draft,userprofile_Story.displayed as displayed,userprofile_Story.id as story_id  FROM api_user LEFT JOIN userprofile_alumni ON api_user.id=userprofile_alumni.user_id LEFT JOIN userprofile_Story ON userprofile_alumni.id=userprofile_Story.alumn_id WHERE api_user.is_alumni=true;")

            # if there is something in items else raise error
            if story:
                serializer = DisplayAllStoriesSerializer(story, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


class StoryHomeView(APIView):
    def get(self, request):
        try:
            # checking for the parameters from the URL
            story = User.objects.raw(
                    "SELECT api_user.id as id, api_user.email as email, api_user.phone1 as phone1, api_user.first_name as first_name, api_user.last_name as last_name,api_user.image_url,userprofile_Story.description as description,userprofile_Story.image,userprofile_Story.title,userprofile_Story.video, userprofile_Story.draft,userprofile_Story.displayed as displayed,userprofile_Story.id as story_id  FROM api_user LEFT JOIN userprofile_alumni ON api_user.id=userprofile_alumni.user_id LEFT JOIN userprofile_Story ON userprofile_alumni.id=userprofile_Story.alumn_id WHERE api_user.is_alumni=true and userprofile_Story.displayed=true;")
            # if there is something in items else raise error
            if story:
                serializer = DisplayAllStoriesSerializer(story, many=True)
                return Response(serializer.data)
            else:
                return Response([])
            
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
# update story is meant for updating the content of the story

def update_story(request, pk):
    story = Story.objects.get(pk=pk)
    data = UpdateStorySerializer(instance=story, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
# display story is meant for displaying the story in the front page

def display_story(request, pk):
    try:
        story = Story.objects.get(pk=pk)
        data = DisplayStorySerializer(instance=story, data=request.data)

        if data.is_valid():
            data.save()
            return Response(data.data)
        else:
                return Response([])
    
    except Exception as e:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_story(request, pk):
    story = get_object_or_404(Story, pk=pk)
    story.delete()
    return Response(status=status.HTTP_202_ACCEPTED)

#new way of handling stories
class StoryViewSet(viewsets.ModelViewSet):
    queryset = Story.objects.all()
    serializer_class = StorySerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except serializers.ValidationError as ve:
            return Response({'error': str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except serializers.ValidationError as ve:
            return Response({'error': str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# end

# Employment view
class EmploymentView(APIView):
    #permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = EmploymentSerializer(data=request.data)
        # validating for already existing data
        if Employment.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            print(serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            if request.query_params:
                alumn = Employment.objects.filter(**request.query_params.dict())
                serializer = EmploymentDisplayOneSerializer(alumn, many=True)
                return Response(serializer.data)
            else:
                user = User.objects.raw("SELECT api_user.id as id,userprofile_alumni.id as alumn_id, api_user.email as email, api_user.phone1 as phone1, api_user.first_name as first_name, api_user.last_name as last_name,api_user.image_url,userprofile_alumni.reg_number,userprofile_employment.title as title,userprofile_employment.company as company,userprofile_employment.description,userprofile_employment.start_date,userprofile_employment.end_date as end,userprofile_employment.status as status,userprofile_employment.id as emp_id,userprofile_employment.career as career,userprofile_family.family_name,userprofile_grade.grade_name,userprofile_combination.combination_name  FROM api_user LEFT JOIN userprofile_alumni ON api_user.id=userprofile_alumni.user_id LEFT JOIN userprofile_family on userprofile_alumni.family_id=userprofile_family.id LEFT JOIN userprofile_grade on userprofile_family.grade_id=userprofile_grade.id LEFT JOIN userprofile_combination on userprofile_alumni.combination_id=userprofile_combination.id  LEFT JOIN userprofile_employment ON userprofile_alumni.id=userprofile_employment.alumn_id WHERE api_user.is_alumni=true;")
                
            # if there is something in items else raise error
            if user:
                serializer = DisplayEmploymentSerializer(user, many=True)
                
                return Response(serializer.data)
            else:
                return Response(["jhjgj"])

        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_Employment(request, pk):
    employment = Employment.objects.get(pk=pk)
    data = EmploymentUpdateSerializer(instance=employment, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_employment(request, pk):
    employment = get_object_or_404(Employment, pk=pk)
    employment.delete()
    return Response(status=status.HTTP_202_ACCEPTED)
class EmploymentBulkCreateUpdateView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            serializer = EmploymentSerializer(data=request.data, many=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request, *args, **kwargs):
        try:
            employment_list = request.data
            for employment_data in employment_list:
                employment_instance = Employment.objects.get(id=employment_data['id'])
                serializer = EmploymentSerializer(employment_instance, data=employment_data)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            return Response({"detail": "Bulk update successful"}, status=status.HTTP_200_OK)
        except Employment.DoesNotExist as e:
            return Response({'detail': str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#Donation track
class SampleMoMoCodeViewSet(viewsets.ModelViewSet):
    queryset = SampleMoMoCode.objects.all()
    serializer_class = SampleMoMoCodeSerializer

class SampleDonationViewSet(viewsets.ModelViewSet):
    queryset = SampleDonation.objects.all()
    serializer_class = SampleDonationSerializer

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except SampleDonation.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)       
#Mentorship program
class MentorshipCardViewSet(viewsets.ModelViewSet):
    queryset = MentorshipCard.objects.all()
    serializer_class = MentorshipCardSerializer

    def handle_exception(self, exc):
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return self.handle_exception(e)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return self.handle_exception(e)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return self.handle_exception(e)

class SampleApplicationsDataViewSet(viewsets.ModelViewSet):
    queryset = SampleApplicationsData.objects.all()
    serializer_class = SampleApplicationsDataSerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            # Handle user ID
            user_id = request.data.get('user_id')
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                    serializer.validated_data['user'] = user
                except User.DoesNotExist:
                    return Response({"error": "User not found"}, status=status.HTTP_400_BAD_REQUEST)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except ValidationError as e:
            return Response({"error": e.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            # Handle user ID
            user_id = request.data.get('user_id')
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                    serializer.validated_data['user'] = user
                except User.DoesNotExist:
                    return Response({"error": "User not found"}, status=status.HTTP_400_BAD_REQUEST)
            self.perform_update(serializer)
            return Response(serializer.data)
        except ValidationError as e:
            return Response({"error": e.detail}, status=status.HTTP_400_BAD_REQUEST)
        except NotFound as e:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except NotFound as e:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# Studie data view

class StudieView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def post(self, request):
        serializer = StudieSerializer(data=request.data)
        # validating for already existing data
        if Studie.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                stud1 = Studie.objects.filter(**request.query_params.dict())
                serializer = StudyWithAlumnSerializer(stud1, many=True)
                return Response(serializer.data)
            else:
                stud = User.objects.raw("SELECT api_user.id as id, api_user.email as email, api_user.phone1 as phone1, api_user.first_name as first_name, api_user.last_name as last_name,api_user.image_url,userprofile_alumni.id as alumn_id,userprofile_alumni.reg_number,userprofile_studie.level,userprofile_studie.degree,userprofile_studie.university,userprofile_studie.country,userprofile_studie.scholarship,userprofile_studie.status,userprofile_studie.id as study_id,userprofile_studie.scholarship_details  FROM api_user LEFT JOIN userprofile_alumni ON api_user.id=userprofile_alumni.user_id LEFT JOIN userprofile_studie ON userprofile_alumni.id=userprofile_studie.alumn_id WHERE api_user.is_alumni=true;")
        
            # if there is something in items else raise error
            if stud:
                serializer = StudieWithAlumnSerializer(stud, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_studie(request, pk):
    stud = Studie.objects.get(pk=pk)
    data = UpdateStudieSerializer(instance=stud, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
#@permission_classes([IsAuthenticated])
def delete_studie(request, pk):
    stud = get_object_or_404(Studie, pk=pk)
    stud.delete()
    return Response(status=status.HTTP_202_ACCEPTED)

@api_view(['POST'])
def bulk_create_studies(request):
    if isinstance(request.data, list):
        serializer = StudieSerializer(data=request.data, many=True)
    else:
        return Response({"error": "Expected a list of items."}, status=status.HTTP_400_BAD_REQUEST)
    
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class BulkUpdateStudieView(APIView):
    def put(self, request, *args, **kwargs):
        data = request.data
        if not isinstance(data, list):
            return Response({"error": "Expected a list of items"}, status=status.HTTP_400_BAD_REQUEST)

        instances = []
        for item in data:
            instance = Studie.objects.get(id=item['id'])
            serializer = StudieSerializer(instance, data=item, partial=True)
            if serializer.is_valid():
                instances.append(serializer.save())
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({"success": "Bulk update successful"}, status=status.HTTP_200_OK)
# end


""" # Opportunity model
@api_view(['GET'])  # GET request, return all opportunities
def read_opportunity(request):
    opportunities = Opportunity.objects.all()
    serializer = OpportunitySerializer(opportunities, many=True)
    return Response(serializer.data)


@api_view(['POST'])  # POST request, create opportunity object
def create_opportunity(request):
    request.data['approved'] = False  # default approved is False
    serializer = OpportunitySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400) """


# Dashboard needed data view


class StudyReportView(APIView):
    permission_classes = [IsAuthenticated, ]  
    def get(self,request):
        try:
            stud = Studie.objects.raw("select 1 as id, count(userprofile_studie.level) as level, userprofile_studie.level as degree from userprofile_studie group by userprofile_studie.level;")
    
            # if there is something in items else raise error
            if stud:
                serializer = StudyReportSerializer(stud, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


# Opportunity model
@api_view(['GET'])
#@permission_classes([IsAuthenticated])
def read_opportunity(request):
    try:
        opportunities = Opportunity.objects.all()
        serializer = OpportunitySerializer(opportunities, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response(e)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_opportunity(request):
    serializer = OpportunitySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


class DeleteOpportunityView(APIView):
    @permission_classes([IsAuthenticated])
    def delete(self, request, pk):
        try:
            opportunity = Opportunity.objects.get(pk=pk)
        except Opportunity.DoesNotExist:
            return Response({'msg': 'Opportunity not found'}, status=status.HTTP_404_NOT_FOUND)

        opportunity.delete()
        return Response({'msg': 'Opportunity deleted successfully'}, status=status.HTTP_200_OK)


class UpdateOpportunityView(RetrieveUpdateAPIView):
    queryset = Opportunity.objects.all()
    serializer_class = UpdateOpportunitySerializer
    lookup_field = 'pk'

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        # check is user is alumni
        if request.user.is_authenticated and request.user.is_alumni:
            raise PermissionDenied("Opportunity has been approved so cannot modify.")

        # update
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)


class ApproveOpportunityView(RetrieveUpdateAPIView):
    queryset = Opportunity.objects.all()
    serializer_class = ApproveOpportunitySerializer
    lookup_field = 'pk'


# Dashboard needed data view

class AlumnReportView(APIView):
    permission_classes = [IsAuthenticated, ]

    def get(self, request):
        try:
            stud = User.objects.raw(
            "SELECT api_user.id,userprofile_alumni.gender,userprofile_grade.grade_name,userprofile_family.family_name,userprofile_combination.combination_name,userprofile_employment.status as employed,userprofile_employment.end_date as end,userprofile_studie.level as degree from api_user left outer join userprofile_alumni on api_user.id=userprofile_alumni.user_id LEFT OUTER JOIN userprofile_family ON userprofile_alumni.family_id=userprofile_family.id left outer join userprofile_grade on userprofile_family.grade_id=userprofile_grade.id LEFT OUTER JOIN userprofile_employment ON userprofile_alumni.id=userprofile_employment.alumn_id LEFT OUTER JOIN userprofile_studie ON userprofile_alumni.id=userprofile_studie.alumn_id left outer join userprofile_combination on userprofile_alumni.combination_id=userprofile_combination.id  WHERE api_user.is_alumni;")

            # if there is something in items else raise error
            if stud:
                serializer = TotalAlumnReportSerializer(stud, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(error=e,status=status.HTTP_404_NOT_FOUND)

class AlumnInGradeReportView(APIView):
    permission_classes = [IsAuthenticated, ]  
    def get(self,request):
        try:
            stud = User.objects.raw("SELECT userprofile_grade.id,  userprofile_grade.grade_name as grade,userprofile_alumni.gender,count(*) as number from api_user left outer join userprofile_alumni on api_user.id=userprofile_alumni.user_id LEFT OUTER JOIN userprofile_family ON userprofile_alumni.Family_id=userprofile_family.id LEFT OUTER JOIN userprofile_grade ON userprofile_family.grade_id=userprofile_grade.id WHERE api_user.is_alumni group by userprofile_grade.id, userprofile_alumni.gender;")
        
            # if there is something in items else raise error
            if stud:
                serializer = TotalAlumnGradeSerializer(stud, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        

class EmploymentStudieReportView(APIView):
    permission_classes = [IsAuthenticated, ]  
    def get(self,request):
        try: 
            empstud = User.objects.raw("SELECT alumni.gender,grade.grade_name,employment.status as emp,studie.level as stu,alumni.id FROM userprofile_alumni AS alumni LEFT JOIN userprofile_employment AS employment ON employment.alumn_id = alumni.id LEFT JOIN userprofile_studie AS studie ON studie.alumn_id=alumni.id JOIN userprofile_family AS family ON alumni.family_id= family.id JOIN userprofile_grade AS grade ON family.grade_id=grade.id;")
            
            # if there is something in items else raise error
            if empstud:
                serializer = EmploymentAndStudieSerializer(empstud, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response( status=status.HTTP_404_NOT_FOUND)
        


# Gallery data view
class GalleryViewSet(viewsets.ModelViewSet):
    queryset = Gallery.objects.all()
    serializer_class = GallerySerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAuthenticated]
        return super().get_permissions()

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Gallery.DoesNotExist:
            return Response({'error': 'Gallery not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except Gallery.DoesNotExist:
            return Response({'error': 'Gallery not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Gallery.DoesNotExist:
            return Response({'error': 'Gallery not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# end


# News
@api_view(['POST'])
def create_news(request):
    serializer = NewsSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


class newsView(APIView):
    #permission_classes = [IsAuthenticated, ]
    
    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                news = News.objects.filter(**request.query_params.dict())
            else:
                news = News.objects.all()

            # if there is something in items else raise error
            if news:
                serializer = NewsSerializer(news, many=True)
                return Response(serializer.data)
            else:
                return Response([])
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)

""" @api_view(['GET'])
def news_list(request):
    news = News.objects.all()
    serializer = NewsSerializer(news, many=True)
    return Response(serializer.data) """


@api_view(['PUT'])
def update_news(request, pk):
    try:
        news = News.objects.get(pk=pk)
    except News.DoesNotExist:
        return Response(status=404)

    if request.data.get('pinned', False):  # if the news is to be pinned
        pinned_news_count = News.objects.filter(pinned=True).count()
        if pinned_news_count > 4:  # if there are already 4 pinned news
            return Response({'error': 'Cannot pin more than 4 news.'}, status=400)  # reject the request

    serializer = NewsSerializer(news, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=400)


@api_view(['DELETE'])
def delete_news(request, pk):
    try:
        news = News.objects.get(pk=pk)
    except News.DoesNotExist:
        return Response(status=404)

    news.delete()
    return Response(status=204)

#news in pdf
class PDFNewsViewSet(viewsets.ModelViewSet):
    queryset = PDFNews.objects.all()
    serializer_class = PDFNewsSerializer

    def create(self, request, *args, **kwargs):
        try:
            # Check if the file is provided in the request
            pdf_file = request.FILES.get('pdf_file', None)
            if not pdf_file:
                raise ValidationError("No file was uploaded.")
            
            # Validate the file type
            if pdf_file.content_type != 'application/pdf':
                raise ValidationError("Only PDF files are allowed.")
            
            # Proceed with the usual creation process if validation passes
            return super().create(request, *args, **kwargs)
        
        except ValidationError as e:
            # Return a response with a 400 Bad Request status
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            # Catch any other unexpected errors
            return Response({'error': 'An unexpected error occurred: ' + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            # Check if the file is provided in the request for an update
            pdf_file = request.FILES.get('pdf_file', None)
            if pdf_file and pdf_file.content_type != 'application/pdf':
                raise ValidationError("Only PDF files are allowed.")
            
            # Proceed with the usual update process if validation passes
            return super().update(request, *args, **kwargs)
        
        except ValidationError as e:
            # Return a response with a 400 Bad Request status
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            # Catch any other unexpected errors
            return Response({'error': 'An unexpected error occurred: ' + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Count alumni to show in front page
def alumni_count(request):
    try:
        count = Alumni.objects.count()  # count Alumni number
        return JsonResponse({"count": count})  # response JSON
    except Exception as e:
        return Response(e)

class UserCountAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        with connection.cursor() as cursor:
            cursor.execute("""
            select COUNT(api_user.id) as total_users,
            COUNT(CASE WHEN userprofile_alumni.gender = 'Male' THEN 1 END) AS male_count,
            COUNT(CASE WHEN userprofile_alumni.gender = 'Female' THEN 1 END) AS female_count 
            from api_user left join userprofile_alumni on api_user.id=userprofile_alumni.user_id 
            where api_user.is_alumni;
            """)
            row = cursor.fetchone()
            
            #for education
        sql_query1 = "select alumn_id,level, CASE WHEN level = 'PHD' THEN 1 WHEN level = 'M' THEN 2 WHEN level= 'A0' THEN 3 WHEN level = 'A1' THEN 4 WHEN level = 'C' THEN 5 WHEN level = 'NMS' THEN 6 WHEN level= 'D' THEN 0 ELSE 7 END AS level_code from userprofile_studie order by alumn_id,level_code desc;"

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            columns = [col[0] for col in cursor1.description]
            data1 = cursor1.fetchall()

        # Create a Pandas DataFrame
        df = pd.DataFrame(data1, columns=columns)
        # Count alumn_id group by level, and for repeated alumn_id, count the one with less level_code
        result_df = df.groupby('alumn_id').apply(lambda group: group.loc[group['level_code'].idxmin()]).groupby('level').agg({'alumn_id': 'count'}).reset_index()
        
        # Rename columns as needed
        result_df.columns = ['level', 'count']
        result_dict = dict(zip(result_df['level'], result_df['count']))
        
        #for employment
        sql_query2 = "select alumn_id,status, CASE WHEN status = 'S' THEN 1 WHEN status = 'F' THEN 2 WHEN status= 'P' THEN 3 WHEN status = 'I' THEN 4 WHEN status = 'U' THEN 5 WHEN status= 'D' THEN 0 ELSE 6 END AS status_code from userprofile_employment order by alumn_id,status_code desc;"

        # Execute the SQL query
        with connection.cursor() as cursor2:
            cursor2.execute(sql_query2)
            columns2 = [col[0] for col in cursor2.description]
            data2 = cursor2.fetchall()

        # Create a Pandas DataFrame
        df2 = pd.DataFrame(data2, columns=columns2)
        
        # Count alumn_id group by status, and for repeated alumn_id, count the one with less status_code
        result_df2 = df2.groupby('alumn_id').apply(lambda group: group.loc[group['status_code'].idxmin()]).groupby('status').agg({'alumn_id': 'count'}).reset_index()
        
        # Rename columns as needed
        result_df2.columns = ['status', 'count']
        
        result_dict2 = dict(zip(result_df2['status'], result_df2['count']))
        
        # Check if the old key exists in the dictionary
        if 'D' in result_dict2:
            # Create a new key with the desired name and assign the value of the old key
            result_dict2['DEM'] = result_dict2.pop('D')
            
        # Check if the old key exists in the dictionary
        if 'N' in result_dict2:
            # Create a new key with the desired name and assign the value of the old key
            result_dict2['NEM'] = result_dict2.pop('N')
        
        
        if row is not None:
            data = {
                'total_users': row[0],
                'male_count': row[1],
                'female_count': row[2]
            }
            data.update(result_dict)
            
            keys=['C','A1','A0','M','PHD','NMS','D','N']
            difference_list = list(set(keys) - set(list(result_dict.keys())))
            
            if(len(difference_list)>0):
                data.update({key: 0 for key in difference_list})
                
            data.update(result_dict2)
            keys2=['S','F','P','I','U','DEM','NEM']
            difference_list2 = list(set(keys2) - set(list(result_dict2.keys())))
            
            if(len(difference_list2)>0):
                data.update({key: 0 for key in difference_list2})

            serializer = AlumniCountSerializer(data)
            return Response(serializer.data)
        else:
            # Handle the case when no rows are returned
            data = {
                'total_users': 0,
                'male_count': 0,
                'female_count': 0,
                'C':0,
                'A1':0,
                'A0':0,
                'M':0,
                'PHD':0,
                'NMS':0,
                'D':0,
                'N':0,
                'F':0,
                'P':0,
                'S':0,
                'I':0,
                'U':0,
                'DEM':0,
                'NEM':0
            }

            serializer = AlumniCountSerializer(data)
            return Response(serializer.data)
        
class UserCountByGradeAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
            
            #count alumni by grade
        sql_query1 = "select userprofile_grade.grade_name,userprofile_grade.end_academic_year, sum(CASE WHEN userprofile_alumni.gender = 'Male' THEN 1 ELSE 0 END) AS male,sum(CASE WHEN userprofile_alumni.gender='Female' THEN 1 ELSE 0 END) AS female from userprofile_alumni inner join userprofile_family on userprofile_alumni.family_id=userprofile_family.id inner join userprofile_grade on userprofile_family.grade_id=userprofile_grade.id group by userprofile_grade.grade_name,userprofile_grade.end_academic_year  order by userprofile_grade.end_academic_year;"

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data=[]   
        if data1 is not None:
            for i in data1:
                data.append({
                        'grade_name': i[0],
                        'male': i[2],
                        'female': i[3]
                })

        serializer = AlumniCountByGradeSerializer(data, many=True)
        return Response(serializer.data)
    
#Count user by combination and gender
class UserCountByCombinationAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
            
            #count alumni by grade
        sql_query1 = "select combination_name, sum(CASE WHEN userprofile_alumni.gender = 'Male' THEN 1 ELSE 0 END) AS male,sum(CASE WHEN userprofile_alumni.gender='Female' THEN 1 ELSE 0 END) AS female from userprofile_alumni inner join userprofile_combination on userprofile_alumni.combination_id=userprofile_combination.id group by userprofile_combination.combination_name order by userprofile_combination.combination_name;"

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data=[]   
        if data1 is not None:
            for i in data1:
                data.append({
                        'combination_name': i[0],
                        'male': i[1],
                        'female': i[2],
                        'total':i[1]+i[2]
                })

        serializer = AlumniCountByCombinationSerializer(data, many=True)
        return Response(serializer.data)
    
class EmploymentStatusByGradeAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
            
            #count alumni by grade
        sql_query1 = """
            WITH alumni_employment AS (
                SELECT DISTINCT ON (a.id)
                    a.id,
                    g.grade_name,
                    g.end_academic_year,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'D'  THEN 1 ELSE 0 END AS diedmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status= 'D' THEN 1 ELSE 0 END AS diedfemale,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'A'  AND e.status IN ('F', 'S', 'P', 'I') THEN 1 ELSE 0 END AS empmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status='A'  AND e.status IN ('F', 'S', 'P', 'I') THEN 1 ELSE 0 END AS empfemale,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'A'  AND (e.status IS NULL OR e.status NOT IN ('F', 'S', 'P', 'I')) THEN 1 ELSE 0 END AS unempmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status= 'A'  AND (e.status IS NULL OR e.status NOT IN ('F', 'S', 'P', 'I')) THEN 1 ELSE 0 END AS unempfemale
                FROM userprofile_alumni a
                INNER JOIN userprofile_family f ON a.family_id = f.id
                INNER JOIN userprofile_grade g ON f.grade_id = g.id
                LEFT JOIN userprofile_employment e ON a.id = e.alumn_id
                ORDER BY a.id,
                    CASE
                        WHEN e.status = 'D' THEN 1
                        WHEN e.status = 'S' THEN 2
                        WHEN e.status = 'F' THEN 3
                        WHEN e.status = 'P' THEN 4
                        WHEN e.status = 'I' THEN 5
                        WHEN e.status = 'U' THEN 6
                        ELSE 7
                    END
            )
            SELECT
                grade_name,
                end_academic_year,
                SUM(diedmale) AS diedmale,
                SUM(diedfemale) AS diedfemale,
                SUM(empmale) AS empmale,
                SUM(empfemale) AS empfemale,
                SUM(unempmale) AS unempmale,
                SUM(unempfemale) AS unempfemale
            FROM alumni_employment
            GROUP BY grade_name, end_academic_year
            ORDER BY end_academic_year;

        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data=[]   
        if data1 is not None:
            for i in data1:
                data.append({
                        'grade_name': i[0],
                        'diedmale': i[2],
                        'diedfemale': i[3],
                        'empmale': i[4],
                        'empfemale': i[5],
                        'unempmale': i[6],
                        'unempfemale': i[7]
                })

        serializer = EmploymentStatusByGradeSerializer(data, many=True)
        return Response(serializer.data)
    
#Employment by families
class EmploymentStatusByFamilyAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
            
            #count alumni by grade
        sql_query1 = """
            WITH alumni_employment AS (
                SELECT DISTINCT ON (a.id)
                    a.id,
                    f.family_name,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'D'  THEN 1 ELSE 0 END AS diedmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status= 'D' THEN 1 ELSE 0 END AS diedfemale,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'A'  AND e.status IN ('F', 'S', 'P', 'I') THEN 1 ELSE 0 END AS empmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status='A'  AND e.status IN ('F', 'S', 'P', 'I') THEN 1 ELSE 0 END AS empfemale,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'A'  AND (e.status IS NULL OR e.status NOT IN ('F', 'S', 'P', 'I')) THEN 1 ELSE 0 END AS unempmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status= 'A'  AND (e.status IS NULL OR e.status NOT IN ('F', 'S', 'P', 'I')) THEN 1 ELSE 0 END AS unempfemale
                FROM userprofile_alumni a
                INNER JOIN userprofile_family f ON a.family_id = f.id
                LEFT JOIN userprofile_employment e ON a.id = e.alumn_id
                ORDER BY a.id,
                    CASE
                        WHEN e.status = 'D' THEN 1
                        WHEN e.status = 'S' THEN 2
                        WHEN e.status = 'F' THEN 3
                        WHEN e.status = 'P' THEN 4
                        WHEN e.status = 'I' THEN 5
                        WHEN e.status = 'U' THEN 6
                        ELSE 7
                    END
            )
            SELECT
                family_name,
                SUM(diedmale) AS diedmale,
                SUM(diedfemale) AS diedfemale,
                SUM(empmale) AS empmale,
                SUM(empfemale) AS empfemale,
                SUM(unempmale) AS unempmale,
                SUM(unempfemale) AS unempfemale
            FROM alumni_employment
            GROUP BY family_name
            ORDER BY family_name;

        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data=[]   
        if data1 is not None:
            for i in data1:
                data.append({
                        'family_name': i[0],
                        'diedmale': i[1],
                        'diedfemale': i[2],
                        'empmale': i[3],
                        'empfemale': i[4],
                        'unempmale': i[5],
                        'unempfemale': i[6]
                })

        serializer = EmploymentStatusByFamilySerializer(data, many=True)
        return Response(serializer.data)
    
#Employment by combinations
class EmploymentStatusByCombinationAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
            
            #count alumni by grade
        sql_query1 = """
            WITH alumni_employment AS (
                SELECT DISTINCT ON (a.id)
                    a.id,
                    c.combination_name,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'D'  THEN 1 ELSE 0 END AS diedmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status= 'D' THEN 1 ELSE 0 END AS diedfemale,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'A'  AND e.status IN ('F', 'S', 'P', 'I') THEN 1 ELSE 0 END AS empmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status='A'  AND e.status IN ('F', 'S', 'P', 'I') THEN 1 ELSE 0 END AS empfemale,
                    CASE WHEN a.gender = 'Male' AND a.life_status= 'A'  AND (e.status IS NULL OR e.status NOT IN ('F', 'S', 'P', 'I')) THEN 1 ELSE 0 END AS unempmale,
                    CASE WHEN a.gender = 'Female' AND a.life_status= 'A'  AND (e.status IS NULL OR e.status NOT IN ('F', 'S', 'P', 'I')) THEN 1 ELSE 0 END AS unempfemale
                FROM userprofile_alumni a
                INNER JOIN userprofile_combination c ON a.combination_id = c.id
                LEFT JOIN userprofile_employment e ON a.id = e.alumn_id
                ORDER BY a.id,
                    CASE
                        WHEN e.status = 'D' THEN 1
                        WHEN e.status = 'S' THEN 2
                        WHEN e.status = 'F' THEN 3
                        WHEN e.status = 'P' THEN 4
                        WHEN e.status = 'I' THEN 5
                        WHEN e.status = 'U' THEN 6
                        ELSE 7
                    END
            )
            SELECT
                combination_name,
                SUM(diedmale) AS diedmale,
                SUM(diedfemale) AS diedfemale,
                SUM(empmale) AS empmale,
                SUM(empfemale) AS empfemale,
                SUM(unempmale) AS unempmale,
                SUM(unempfemale) AS unempfemale
            FROM alumni_employment
            GROUP BY combination_name
            ORDER BY combination_name;

        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data=[]   
        if data1 is not None:
            for i in data1:
                data.append({
                        'family_name': i[0],
                        'diedmale': i[1],
                        'diedfemale': i[2],
                        'empmale': i[3],
                        'empfemale': i[4],
                        'unempmale': i[5],
                        'unempfemale': i[6]
                })

        serializer = EmploymentStatusByFamilySerializer(data, many=True)
        return Response(serializer.data)
    
class StudieStatusByGradeAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
        # count alumni by grade
        sql_query1 = """
            SELECT 
                grade_name,
                end_academic_year,
                SUM(diedmale) AS diedmale,
                SUM(diedfemale) AS diedfemale, 
                SUM(stumale) AS stumale,
                SUM(stufemale) AS stufemale,
                SUM(nstumale) AS nstumale, 
                SUM(nstufemale) AS nstufemale
            FROM (
                SELECT DISTINCT ON (u.id) 
                    g.grade_name,
                    g.end_academic_year,
                    CASE 
                        WHEN u.gender='Male' AND u.life_status='D' THEN 1 
                        ELSE 0 
                    END AS diedmale,
                    CASE 
                        WHEN u.gender='Female' AND u.life_status='D' THEN 1 
                        ELSE 0 
                    END AS diedfemale,
                    CASE 
                        WHEN u.gender='Male' AND s.level IN ('C','A1','A0','M','PHD') AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS stumale,
                    CASE 
                        WHEN u.gender='Female' AND s.level IN ('C','A1','A0','M','PHD') AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS stufemale,
                    CASE 
                        WHEN u.gender='Male' AND (s.level NOT IN ('C','A1','A0','M','PHD') OR s.level IS NULL) AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS nstumale,
                    CASE 
                        WHEN u.gender='Female' AND (s.level NOT IN ('C','A1','A0','M','PHD') OR s.level IS NULL) AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS nstufemale
                FROM userprofile_alumni u
                INNER JOIN userprofile_family f ON u.family_id = f.id
                INNER JOIN userprofile_grade g ON f.grade_id = g.id
                LEFT JOIN userprofile_studie s ON u.id = s.alumn_id
                ORDER BY u.id, 
                    CASE 
                        WHEN s.level = 'D' THEN 1 
                        WHEN s.level = 'PHD' THEN 2 
                        WHEN s.level= 'M' THEN 3 
                        WHEN s.level= 'A0' THEN 4 
                        WHEN s.level= 'A1' THEN 5 
                        WHEN s.level= 'C' THEN 6 
                        WHEN s.level= 'NMS' THEN 7 
                        ELSE 8 
                    END
            ) AS studie
            GROUP BY grade_name, end_academic_year 
            ORDER BY end_academic_year;
        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data = []
        if data1 is not None:
            for i in data1:
                data.append({
                    'grade_name': i[0],
                    'end_academic_year': i[1],  # include end_academic_year in the response
                    'diedmale': i[2],
                    'diedfemale': i[3],
                    'stumale': i[4],
                    'stufemale': i[5],
                    'nstumale': i[6],
                    'nstufemale': i[7]
                })

        serializer = StudieStatusByGradeSerializer(data, many=True)
        return Response(serializer.data)
    
#Studies by families
class StudieStatusByFamilyAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
        # count alumni by family
        sql_query1 = """
            SELECT 
                family_name,
                SUM(diedmale) AS diedmale,
                SUM(diedfemale) AS diedfemale, 
                SUM(stumale) AS stumale,
                SUM(stufemale) AS stufemale,
                SUM(nstumale) AS nstumale, 
                SUM(nstufemale) AS nstufemale
            FROM (
                SELECT DISTINCT ON (u.id) 
                    f.family_name,
                    CASE 
                        WHEN u.gender='Male' AND u.life_status='D' THEN 1 
                        ELSE 0 
                    END AS diedmale,
                    CASE 
                        WHEN u.gender='Female' AND u.life_status='D' THEN 1 
                        ELSE 0 
                    END AS diedfemale,
                    CASE 
                        WHEN u.gender='Male' AND s.level IN ('C','A1','A0','M','PHD') AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS stumale,
                    CASE 
                        WHEN u.gender='Female' AND s.level IN ('C','A1','A0','M','PHD') AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS stufemale,
                    CASE 
                        WHEN u.gender='Male' AND (s.level NOT IN ('C','A1','A0','M','PHD') OR s.level IS NULL) AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS nstumale,
                    CASE 
                        WHEN u.gender='Female' AND (s.level NOT IN ('C','A1','A0','M','PHD') OR s.level IS NULL) AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS nstufemale
                FROM userprofile_alumni u
                INNER JOIN userprofile_family f ON u.family_id = f.id
                LEFT JOIN userprofile_studie s ON u.id = s.alumn_id
                ORDER BY u.id, 
                    CASE 
                        WHEN s.level = 'D' THEN 1 
                        WHEN s.level = 'PHD' THEN 2 
                        WHEN s.level= 'M' THEN 3 
                        WHEN s.level= 'A0' THEN 4 
                        WHEN s.level= 'A1' THEN 5 
                        WHEN s.level= 'C' THEN 6 
                        WHEN s.level= 'NMS' THEN 7 
                        ELSE 8 
                    END
            ) AS studie
            GROUP BY family_name
            ORDER BY family_name;
        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data = []
        if data1 is not None:
            for i in data1:
                data.append({
                    'family_name': i[0],
                    'diedmale': i[1],
                    'diedfemale': i[2],
                    'stumale': i[3],
                    'stufemale': i[4],
                    'nstumale': i[5],
                    'nstufemale': i[6]
                })

        serializer = StudieStatusByFamilySerializer(data, many=True)
        return Response(serializer.data)
    
#Studies by combinations
class StudieStatusByCombinationsAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
        
        # count alumni by combination
        sql_query1 = """
            SELECT 
                combination_name,
                SUM(diedmale) AS diedmale,
                SUM(diedfemale) AS diedfemale, 
                SUM(stumale) AS stumale,
                SUM(stufemale) AS stufemale,
                SUM(nstumale) AS nstumale, 
                SUM(nstufemale) AS nstufemale
            FROM (
                SELECT DISTINCT ON (u.id) 
                    c.combination_name,
                    CASE 
                        WHEN u.gender='Male' AND u.life_status='D' THEN 1 
                        ELSE 0 
                    END AS diedmale,
                    CASE 
                        WHEN u.gender='Female' AND u.life_status='D' THEN 1 
                        ELSE 0 
                    END AS diedfemale,
                    CASE 
                        WHEN u.gender='Male' AND s.level IN ('C','A1','A0','M','PHD') AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS stumale,
                    CASE 
                        WHEN u.gender='Female' AND s.level IN ('C','A1','A0','M','PHD') AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS stufemale,
                    CASE 
                        WHEN u.gender='Male' AND (s.level NOT IN ('C','A1','A0','M','PHD') OR s.level IS NULL) AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS nstumale,
                    CASE 
                        WHEN u.gender='Female' AND (s.level NOT IN ('C','A1','A0','M','PHD') OR s.level IS NULL) AND u.life_status='A' THEN 1 
                        ELSE 0 
                    END AS nstufemale
                FROM userprofile_alumni u
                INNER JOIN userprofile_combination c ON u.combination_id = c.id
                LEFT JOIN userprofile_studie s ON u.id = s.alumn_id
                ORDER BY u.id, 
                    CASE 
                        WHEN s.level = 'D' THEN 1 
                        WHEN s.level = 'PHD' THEN 2 
                        WHEN s.level= 'M' THEN 3 
                        WHEN s.level= 'A0' THEN 4 
                        WHEN s.level= 'A1' THEN 5 
                        WHEN s.level= 'C' THEN 6 
                        WHEN s.level= 'NMS' THEN 7 
                        ELSE 8 
                    END
            ) AS studie
            GROUP BY combination_name
            ORDER BY combination_name;
        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data = []
        if data1 is not None:
            for i in data1:
                data.append({
                    'combination_name': i[0],
                    'diedmale': i[1],
                    'diedfemale': i[2],
                    'stumale': i[3],
                    'stufemale': i[4],
                    'nstumale': i[5],
                    'nstufemale': i[6]
                })

        serializer = StudieStatusByCombinationSerializer(data, many=True)
        return Response(serializer.data)


class StudieEmployStatusByGradeAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get(self, request, *args, **kwargs):
            #count alumni by grade
        sql_query1 = """
            SELECT 
                ug.grade_name AS grade, 
                ug.end_academic_year AS ac, 
                SUM(CASE 
                    WHEN ua.gender = 'Male' 
                        AND ue.status IN ('F', 'S', 'P', 'I') 
                        AND us.level IN ('C', 'A1', 'A0', 'M', 'PHD') 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS empstumale, 
                SUM(CASE 
                    WHEN ua.gender = 'Male' 
                        AND ue.status IN ('F', 'S', 'P', 'I') 
                        AND us.level IS NULL 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS empnstumale, 
                SUM(CASE 
                    WHEN ua.gender = 'Female' 
                        AND ue.status IN ('F', 'S', 'P', 'I') 
                        AND us.level IN ('C', 'A1', 'A0', 'M', 'PHD') 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS empstufemale, 
                SUM(CASE 
                    WHEN ua.gender = 'Female' 
                        AND ue.status IN ('F', 'S', 'P', 'I') 
                        AND us.level IS NULL 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS empnstufemale, 
                SUM(CASE 
                    WHEN ua.gender = 'Male' 
                        AND ue.status IS NULL 
                        AND us.level IN ('C', 'A1', 'A0', 'M', 'PHD') 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS unempstumale, 
                SUM(CASE 
                    WHEN ua.gender = 'Male' 
                        AND ue.status IS NULL 
                        AND us.level IS NULL 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS unempnstumale, 
                SUM(CASE 
                    WHEN ua.gender = 'Female' 
                        AND ue.status IS NULL 
                        AND us.level IN ('C', 'A1', 'A0', 'M', 'PHD') 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS unempstufemale, 
                SUM(CASE 
                    WHEN ua.gender = 'Female' 
                        AND ue.status IS NULL 
                        AND us.level IS NULL 
                        AND ua.life_status = 'A'  
                    THEN 1 
                    ELSE 0 
                END) AS unempnstufemale, 
                SUM(CASE 
                    WHEN ua.gender = 'Male'
                        AND ua.life_status = 'D' 
                    THEN 1 
                    ELSE 0 
                END) AS diedumale, 
                SUM(CASE 
                    WHEN ua.gender = 'Female' 
                        AND ua.life_status = 'D' 
                    THEN 1 
                    ELSE 0 
                END) AS diedfemale 
            FROM 
                userprofile_alumni ua 
                INNER JOIN userprofile_family uf ON ua.family_id = uf.id 
                INNER JOIN userprofile_grade ug ON uf.grade_id = ug.id 
                LEFT JOIN (
                    SELECT DISTINCT ON (alumn_id) * 
                    FROM userprofile_employment 
                    ORDER BY alumn_id, status DESC
                ) ue ON ua.id = ue.alumn_id 
                LEFT JOIN (
                    SELECT DISTINCT ON (alumn_id) * 
                    FROM userprofile_studie 
                    ORDER BY alumn_id, level DESC
                ) us ON ua.id = us.alumn_id 
            GROUP BY 
                ug.grade_name, 
                ug.end_academic_year 
            ORDER BY 
                ac;

        """

        # Execute the SQL query
        with connection.cursor() as cursor1:
            cursor1.execute(sql_query1)
            data1 = cursor1.fetchall()
            
        data=[]   
        if data1 is not None:
            for i in data1:
                data.append({
                        'grade_name': i[0],
                        'ac': i[1],
                        'empstumale': i[2],
                        'empnstumale': i[3],
                        'empstufemale': i[4],
                        'empnstufemale': i[5],
                        'unempstumale': i[6],
                        'unempnstumale': i[7],
                        'unempstufemale': i[8],
                        'unempnstufemale': i[9],
                        'diedmale': i[10],
                        'diedfemale': i[11]
                })

        serializer = StudieEmployStatusByGradeSerializer(data, many=True)
        return Response(serializer.data)
    
#def create_data_validation(sheet, column_letter, start_row, options):
        # data_validation = DataValidation(type="list", formula1=options, showDropDown=True)

        # for row in range(start_row, sheet.max_row + 1):
        #     cell = sheet[column_letter + str(row)]
        #     sheet.add_data_validation(data_validation)
        #     data_validation.add(cell)
class UsersExcelExportView(APIView):
    
    def get(self, request, *args, **kwargs):

        # Create an Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        # Get the sheet you want to rename
        alumni_profile = workbook['Sheet']  # Replace 'Sheet1' with the current name of your sheet

        # Set the new name for the sheet
        alumni_profile_name = 'personal_profile'
        alumni_profile.title = alumni_profile_name

        # Write headers excluding sensitive fields
        headers = ["email" ,"first_name" ,"last_name","phone1", "other_emails", "other_phones","date_of_birth","gender","reg_number", "father" ,"mother" ,"did_you_born_in_rwanda","place_of_birth_district_or_country","place_of_birth_sector_or_city","life_status" ,"marital_status", "currresidence_in_rwanda" ,"currresidence_district_or_country","currresidence_sector_or_city" ,"kids" , "family", "combination","eps","s4marks","s5marks","s6marks","ne" ,"decision"]
        alumni_profile.append(headers)
        column_letter_for_datavalidation = 'G' 
        start_row = 2
        # Define options for the dropdown list
        options = '"Yes,No"'

        
        # Iterate over all columns and adjust their widths
        for column in alumni_profile.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            alumni_profile.column_dimensions[column_letter].width = adjusted_width
            
        # Create a new sheet
        employement_sheet_name = 'employement'
        employement_sheet = workbook.create_sheet(title=employement_sheet_name)
        # Write headers excluding sensitive fields
        employement_headers = ["email","job_title","job_status","career","company","current_old"]
        employement_sheet.append(employement_headers)
        
        # Iterate over all columns and adjust their widths
        for column in employement_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            employement_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create a new sheet
        study_sheet_name = 'futher_study'
        study_sheet = workbook.create_sheet(title=study_sheet_name)
        # Write headers excluding sensitive fields
        study_sheet_headers = ["email","study_level","course_name","university","in_which_country","city","Which_scholarship_did_you_receive","Scholarship details (Example: REB, FARG,...)","Study_status"]
        study_sheet.append(study_sheet_headers)
        
        # Iterate over all columns and adjust their widths
        for column in study_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            study_sheet.column_dimensions[column_letter].width = adjusted_width

        try:
            #create_data_validation(alumni_profile, column_letter_for_datavalidation, start_row, options)
            # Create a response with the Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=bulk_alumni_reg_template.xlsx'
            workbook.save(response)
            #print("Data validation applied successfully.")
        except Exception as e:
            print(f"Error: {str(e)}")
        

        return response
    
def is_date(string):
    try:
        # Attempt to parse the string into a date
        parser.parse(str(string))
        return True
    except ValueError:
        # If parsing fails, it's not a date
        return False

def get_family_id(family_name):
    try:
        return Family.objects.get(family_name=family_name).id
    except Family.DoesNotExist:
        # Handle the case where Family with the specified family_name doesn't exist
        return np.nan  # or any default value you prefer
    
def get_combination_id(combination_name):
    try:
        return Combination.objects.get(combination_name=combination_name).id
    except Combination.DoesNotExist:
        # Handle the case where Family with the specified family_name doesn't exist
        return np.nan  # or any default value you prefer

# Function to get alumni ID from email
def get_alumni_id(email):
    try:
        alumni = Alumni.objects.get(user__email=email)
        return alumni.id
    except Alumni.DoesNotExist:
        return np.nan  # Handle the case where the alumni doesn't exist

# Define a function to get eps IDs from the Django model

def get_eps_id(ep):
    try:
        return Ep.objects.get(title=ep).id
    except Ep.DoesNotExist:
        # Handle the case where the episode doesn't exist in the model
        return np.nan
    
def is_excel_file(file):
    # Check if the file has an Excel extension (e.g., '.xlsx' or '.xls')
    allowed_extensions = {'.xlsx', '.xls'}
    _, file_extension = os.path.splitext(file.name)
    
    if file_extension.lower() in allowed_extensions:
        return True
    else:
        return False
class ExcelUploadAPIView(APIView):
    parser_classes = (MultiPartParser,)

    def post(self, request, format=None):
        file = request.FILES.get('file')
        if file:
            if is_excel_file(file):
                print("The uploaded file is an Excel file.")
                
                workbook = openpyxl.load_workbook(file)
                sheet_names = workbook.sheetnames
                data = {}
                expected_sheet_names=['personal_profile','employement','futher_study']
                if set(sheet_names) ==set(expected_sheet_names):
                    for sheet_name in sheet_names:
                        sheet = workbook[sheet_name]
                        rows = []

                        for row in sheet.iter_rows(values_only=True):
                            rows.append(row)

                        data[sheet_name] = rows
                    
                    #...........validate alumni profile data.........................
                    
                    if 'personal_profile' in data:
                        # Extract headers and data
                        alumni_profile_headers = data['personal_profile'][0]
                        if set(data['personal_profile'][0]) == set(('email', 'first_name', 'last_name', 'phone1', 'date_of_birth', 'gender', 'did_you_born_in_rwanda', 'place_of_birth_district_or_country', 'place_of_birth_sector_or_city', 'family', 'combination', 'eps', 's4marks', 's5marks', 's6marks', 'ne', 'maxforne', 'decision', 'life_status', 'marital_status', 'currresidence_in_rwanda', 'currresidence_district_or_country', 'currresidence_sector_or_city', 'kids')):
                            alumni_profile_rows=data['personal_profile'][1:]
                            if(len(alumni_profile_rows)>0):
                                # Create a DataFrame
                                alumni_profile= pd.DataFrame(alumni_profile_rows, columns=alumni_profile_headers)
                                
                                # Check for null values in the entire DataFrame
                                # Sum the null values in each column
                                null_countsalumni_profile = alumni_profile.isnull().sum()
                                if(null_countsalumni_profile>0).all():
                                    print("\nNull counts in each column in profile:")
                                    print(null_countsalumni_profile)
                                    data["error"]="there are some empty cells in profile"
                                    return Response(data)
                                # Check for duplicates in 'email'
                                duplicatesemail = alumni_profile.duplicated('email')
                                # Display rows with duplicates in 'your_column'
                                duplicate_rowsemail = alumni_profile[duplicatesemail]
                                print("Rows with duplicates in email:")
                                print(sum(duplicatesemail))
                                print(duplicate_rowsemail)
                                if(sum(duplicatesemail)>0):
                                    data["error"]="there are some duplicate emails"
                                    return Response(data)
                                
                                #check if email exist
                                existemails=[]
                                for email_to_check in alumni_profile['email']:
                                    if User.objects.filter(email=email_to_check).exists():
                                        existemails.append(email_to_check)
                                  
                                if(len(existemails)>0):
                                    print("List of exist emails:\n")
                                    print(existemails)
                                    data['error']="There are "+str(len(existemails))+" exist emails"
                                    return Response(data)
                                        
                                #check collect email
                                incolect_email=[]
                                # Regular expression for basic email validation
                                email_pattern = r'^\s*[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\s*$'
                                for email_to_check in alumni_profile['email']:
                                    if not (re.match(email_pattern, email_to_check)):
                                        incolect_email.append(email_to_check)
                                        
                                print("List of incollect emails:\n")
                                print(incolect_email)
                                if len(incolect_email)>0:
                                    data["error"]="there are some incorrect emails"
                                    return Response(data)
                                
                                
                                # Check for duplicates in 'phone1'
                                duplicatesphone1 = alumni_profile.duplicated('phone1')
                                # Display rows with duplicates in 'your_column'
                                duplicate_rowsphone1 = alumni_profile[duplicatesphone1]
                                print("Rows with duplicates in phone1:")
                                print(sum(duplicatesphone1))
                                print(duplicate_rowsphone1)
                                if(sum(duplicatesphone1)):
                                    data["error"]="there are some duplicate phones"
                                    return Response(data)
                                
                                #check if phone1 exist
                                existphones=[]
                                for phone_to_check in alumni_profile['phone1']:
                                    if User.objects.filter(phone1=phone_to_check).exists():
                                        existphones.append(phone_to_check)
                                        
                                print("List of exist phones:\n")
                                print(existphones)
                                if(len(existphones)>0):
                                    data["error"]="there are some exist phones"
                                    return Response(data)
                                
                                #check if current date minus date_of_birth >18
                                
                                
                                invalid_date=[]
                                for dt in alumni_profile['date_of_birth']:
                                    if is_date(dt):
                                        continue
                                    else:
                                        invalid_date.append(alumni_profile.loc[alumni_profile['date_of_birth']==dt,['email']]['email'])
                                        

                                if(len(invalid_date)>0):
                                    data["error"]="there are some invalid dates"
                                    return Response(data)
                                #print("Invalid dob")
                                for dte in invalid_date:
                                    print(dte)
                                    
                                if(len(invalid_date)==0):
                                    # Convert 'date_of_birth' to datetime objects
                                    alumni_profile['date_of_births'] = pd.to_datetime(alumni_profile['date_of_birth'])
                                    # Calculate the age by subtracting 'date_of_birth' from the current date
                                    alumni_profile['age'] =  (datetime.now() - alumni_profile['date_of_births']) // pd.Timedelta('365.25D')

                                    # Filter rows where age is less than to 18
                                    filtered_alumni_profile = alumni_profile[alumni_profile['age'] < 18]

                                    # Display the resulting DataFrame with emails
                                    print("People with age bellow 18")
                                    print(filtered_alumni_profile[['email']])
                                    if(len(filtered_alumni_profile)>0):
                                        data["error"]="there some People with age bellow 18"
                                        return Response(data)
                                    
                                    # Check if all values in the 'gender' column are either 'Female' or 'Male'
                                    valid_genders = set(['female', 'male'])
                                    is_valid = alumni_profile['gender'].str.lower().isin(valid_genders).all()

                                    # If all values are valid, capitalize the 'gender' column
                                    if is_valid:
                                        alumni_profile['gender'] = alumni_profile['gender'].str.capitalize()
                                        print("All values in the 'gender' column are either 'Female' or 'Male'. Capitalized the values.")
                                    else:
                                        print("Some values in the 'gender' column are not 'Female' or 'Male'. No changes made.")
                                        data["error"]="Some values in the 'gender' column are not 'Female' or 'Male'. No changes made."
                                        return Response(data)
                                        
                                    # Check if all values in the 'did_you_born_in_rwanda' column are either 'Yes' or 'No'
                                    valid_did_you_born_in_rwanda = set(['yes', 'no'])
                                    is_valid_did_you_born_in_rwanda = alumni_profile['did_you_born_in_rwanda'].str.lower().isin(valid_did_you_born_in_rwanda).all()

                                    # If all values are valid, capitalize the 'did_you_born_in_rwanda' column
                                    if is_valid_did_you_born_in_rwanda:
                                        alumni_profile['did_you_born_in_rwanda'] = alumni_profile['did_you_born_in_rwanda'].str.capitalize()
                                        print("All values in the 'did_you_born_in_rwanda' column are either 'Yes' or 'No'. Capitalized the values.")
                                    else:
                                        print("Some values in the 'did_you_born_in_rwanda' column are not 'Yes' or 'No'. No changes made.")
                                        data["error"]="Some values in the 'did_you_born_in_rwanda' column are not 'Yes' or 'No'. No changes made."
                                        return Response(data)
                                        
                                    #Add family id
                                    
                                    # Iterate through rows and query the Family model to get family ids
                                    alumni_profile['familyid'] = alumni_profile['family'].apply(get_family_id)

                                    # Alternatively, you can use the map function
                                    # alumni_profile['familyid'] = alumni_profile['familyname'].map(lambda x: Family.objects.get(family_name=x).id)

                                    # Display noumber of non existing family
                                    
                                    if(alumni_profile['familyid'].isnull().sum())>0:
                                        print("number of non existing family:"+str(alumni_profile['familyid'].isnull().sum()))
                                        print(alumni_profile['familyid'])
                                        data["error"]="number of non existing family:"+str(alumni_profile['familyid'].isnull().sum())
                                        return Response(data)
                                    
                                    # Iterate through rows and query the Family model to get combination ids
                                    alumni_profile['combinationid'] = alumni_profile['combination'].apply(get_combination_id)
                                    # Display noumber of non existing combination
                                    print("number of non existing combination:"+str(alumni_profile['combinationid'].isnull().sum()))
                                    if (alumni_profile['combinationid'].isnull().sum())>0:
                                        data["error"]="number of non existing combination:"+str(alumni_profile['combinationid'].isnull().sum())
                                        return Response(data)
                                    
                                    # Apply the function to create a new column 'epsid'
                                    alumni_profile['epsid'] = alumni_profile['eps'].apply(lambda eps: [get_eps_id(ep) for ep in eps.split(',')])
                                    # Select rows where 'epsid' column has NaN in the list
                                    selected_rows = alumni_profile[alumni_profile['epsid'].apply(lambda x: any(pd.isna(y) for y in x))]

                                    # Display the selected rows
                                    print(selected_rows)
                                    if len(selected_rows)>0:
                                        data["error"]="There are some non existing eps"
                                        return Response(data)
                                    
                                    print(alumni_profile[['combinationid','combination','family','familyid','eps','epsid']])
                                    
                                
                                    # Specify the columns you want to check
                                    columns_to_check = ['s4marks', 's5marks', 's6marks', 'ne', 'maxforne']

                                    # Check if all values are numbers and within the specified range
                                    is_valid = alumni_profile[columns_to_check].apply(lambda x: x.apply(lambda y: isinstance(y, (int, float)) and 0 <= y <= 100))

                                    # Check if all values in each column meet the conditions
                                    all_columns_valid = is_valid.all(axis=0)

                                    # Display the result
                                    if sum(all_columns_valid)<5:
                                        print(all_columns_valid)
                                        print("There are some marks which are not in (int, float)) and 0 <= y <= 100)")
                                        data["error"]="There are some marks which are not in (int, float)) and 0 <= y <= 100)"
                                        return Response(data)
                                    # Check if values in ne are less than the corresponding values in maxforne
                                    is_less_than_maxforne = alumni_profile['ne'] < alumni_profile['maxforne']

                                    # Display the result
                                    if not (is_less_than_maxforne).all():
                                        print(is_less_than_maxforne)
                                        print("There are some marks in national exam which are greater that maximum")
                                        data["error"]="There are some marks in national exam which are greater that maximum"
                                        return Response(data)
                                    
                                    
                                    # Define expected categories
                                    expected_decision_values = ['Pass', 'Fail']
                                    expected_life_status_values = ['Alive', 'Died']
                                    expected_residence_values = ['Yes', 'No']
                                    expected_kids_values = ['Yes', 'No']
                                    # Check for incorrect data in each column
                                    incorrect_decision = ~alumni_profile['decision'].str.capitalize().isin(expected_decision_values)
                                    incorrect_life_status = ~alumni_profile['life_status'].str.capitalize().isin(expected_life_status_values)
                                    incorrect_residence = ~alumni_profile['currresidence_in_rwanda'].str.capitalize().isin(expected_residence_values)
                                    incorrect_kids = ~alumni_profile['kids'].str.capitalize().isin(expected_kids_values)

                                    # Filter rows with incorrect data
                                    incorrect_rows = alumni_profile[incorrect_decision | incorrect_life_status | incorrect_residence | incorrect_kids]
                                    
                                    # Display the result
                                    if not (incorrect_rows).empty:
                                        print(incorrect_rows)
                                        print("check if  values in decision column fall in Pass or Fail, life_status column fall in Alive or Died, currresidence_in_rwanda column fall in Yes or No and kids column fall in Yes or No")
                                        data["error"]="check if  values in decision column fall in Pass or Fail, life_status column fall in Alive or Died, currresidence_in_rwanda column fall in Yes or No and kids column fall in Yes or No"
                                        return Response(data)
                                    #Data validation on employement
                                    # Extract headers and data
                                    employement_headers = data['employement'][0]
                                    if(set(employement_headers)==set(('email', 'job_title', 'job_status', 'career', 'company'))):
                                        employement_rows=data['employement'][1:]
                                        if(len(alumni_profile_rows)>0):
                                            # Create a DataFrame
                                            employement= pd.DataFrame(employement_rows, columns=employement_headers)
                                            
                                            # Check for null values in the entire DataFrame
                                            # Sum the null values in each column
                                            null_countsemployement = employement.isnull().sum()
                                            if(null_countsemployement>0).all():
                                                print("\nNull counts in each column in employement:")
                                                print(null_countsemployement)
                                                data["error"]="there are some empty cells in employement"
                                                return Response(data)
                                            
                                            
                                            #ad user id according to email given
                                            # Apply the function to create the 'alumniid' column
                                            #employement['alumniid'] = employement['email'].apply(get_alumni_id)
                                            
                                            #Check if all email in employement sheet have their corresponding email in personal_profile sheet
                                            all_emails_present_in_employ = employement['email'].isin(alumni_profile['email']).all()

                                            if all_emails_present_in_employ:
                                                print("All emails in employment have corresponding emails in personal_profile.")
                                            else:
                                                print("Not all emails in employment have corresponding emails in personal_profile.")
                                                # Find emails in employment without correspondence in personal_profile
                                                missing_emails = employement[~employement['email'].isin(alumni_profile['email'])]

                                                print("Emails without correspondence in personal_profile:")
                                                print(missing_emails)
                                                data["error"]="Not all emails in employment have corresponding emails in personal_profile."
                                                return Response(data)
                                           
                                            # Define a mapping dictionary for the first letter to status
                                            status_mapping = {'F': 'F', 'P': 'P', 'S': 'S', 'I': 'I', 'U': 'U', 'D': 'D', 'N': 'N'}
                                            # Create a new 'status' column based on the first letter of 'job_status'
                                            employement['job_status']=employement['job_status'].str.capitalize()
                                            employement['status'] = employement['job_status'].str[0].map(status_mapping)
                                            null_countsemployement2 = employement.isnull().sum()
                                            if(null_countsemployement2>0).all():
                                                print("\nNull counts in each column in employement:")
                                                print(null_countsemployement2)
                                                data["error"]="there are some empty cells in employement especialy in status"
                                                return Response(data)
                                            
                                    #Data validation on further study
                                    # Extract headers and data
                                    futher_study_headers = data['futher_study'][0]
                                    if(set(futher_study_headers)==set(('email', 'study_level', 'course_name', 'university', 'in_which_country', 'Which_scholarship_did_you_receive', 'Scholarship details (Example: REB, FARG,...)', 'Study_status'))):
                                        futher_study_rows=data['futher_study'][1:]
                                        if(len(futher_study_rows)>0):
                                            # Create a DataFrame
                                            futher_study= pd.DataFrame(futher_study_rows, columns=futher_study_headers)
                                            
                                            # Check for null values in the entire DataFrame
                                            # Sum the null values in each column
                                            
                                            
                                            #ad user id according to email given
                                            # Apply the function to create the 'alumniid' column
                                            #futher_study['alumniid'] = futher_study['email'].apply(get_alumni_id)
                                            
                                            #Validate scholarship column
                                            status_mapping_study = {'Full':'F','Partial':'P','Noscholarship':'NS','Deceased':'D','Noinfo':'N','Nofurtherstudy':'NMS'}
                                            # Create a new 'scholar' column based on the first letter of 'Which_scholarship_did_you_receive'
                                            futher_study['Which_scholarship_did_you_receive']=futher_study['Which_scholarship_did_you_receive'].str.capitalize()
                                            futher_study['scholar'] = futher_study['Which_scholarship_did_you_receive'].map(status_mapping_study)
                                            
                                            #Validate study_level column
                                            status_mapping_study_level = {'Certificate':'C','Diploma':'A1','Bachelor':'A0','Masters':'M','Phd':'PHD','Deceased':'D','Noinfo':'N','Nofurtherstudy':'NMS'}
                                            # Create a new 'level' column based on the first letter of 'study_level'
                                            futher_study['study_level']=futher_study['study_level'].str.capitalize()
                                            futher_study['level'] = futher_study['study_level'].map(status_mapping_study_level)
                                            
                                            #Validate study_status column
                                            status_mapping_study_status = {'Graduated':'C','Dropped':'D','Suspended':'S','Studying':'O','Deceased':'D','Noinfo':'N','Nofurtherstudy':'NMS'}
                                            # Create a new 'status' column based on the first letter of 'Study_status'
                                            futher_study['Study_status']=futher_study['Study_status'].str.capitalize()
                                            futher_study['stutas'] = futher_study['Study_status'].map(status_mapping_study_status)
                                            print(futher_study)
                                            """ for index, row in alumni_profile.iterrows():
                                                print(f"Index: {index}, email: {row['email']}") """
                                            
                                            null_countsfuther_study = futher_study.isnull().sum()
                                            if(null_countsemployement2>0).all():
                                                print("\nNull counts in each column in further study:")
                                                print(null_countsfuther_study)
                                                data["error"]="there are some empty cells in further study"
                                                return Response(data)
                                            
                                            #group all data in one dataframe
                                            grouped_employment = employement.groupby('email').agg(lambda x: list(x))
                                            grouped_data = futher_study.groupby('email').agg(lambda x: list(x))
                                            # Merge alumni_profile with grouped_data and grouped_employment
                                            alumni_profile = pd.merge(alumni_profile, grouped_data, on='email', how='left')
                                            alumni_profile = pd.merge(alumni_profile, grouped_employment, on='email', how='left')
                                            
                                            # Add two new columns as dictionaries of related data
                                            #'study_level', 'course_name', 'university', 'in_which_country','Which_scholarship_did_you_receive','Scholarship details (Example: REB, FARG,...)', 'Study_status','alumniid', 'scholar', 'level', 'stutas'
                                            alumni_profile['further_study_data'] = alumni_profile.apply(lambda row: {'study_level': row['level'], 'course_name': row['course_name'], 'university': row['university'], 'in_which_country': row['in_which_country'], 'scholar_details': row['Scholarship details (Example: REB, FARG,...)'], 'scholar': row['scholar'], 'study_status': row['stutas']}, axis=1)

                                            alumni_profile['employment_data'] = alumni_profile.apply(lambda row: {'job_title': row['job_title'], 'job_status': row['job_status'], 'career': row['career'], 'company': row['company'], 'job_status': row['status']}, axis=1)

                                            # Drop unnecessary columns
                                            alumni_profile.drop(['study_level', 'course_name', 'university', 'in_which_country', 'Which_scholarship_did_you_receive', 'Scholarship details (Example: REB, FARG,...)', 'Study_status', 'scholar', 'level', 'stutas', 'job_title', 'job_status', 'career', 'company', 'status'], axis=1, inplace=True)

                                            for index, row in alumni_profile.iterrows():
                                                print(f"Index: {index}, email: {row['email']} First Name: {row['first_name']}")
                                                employdata=row['employment_data']
                                                life = 'A' if row['life_status'].lower().startswith('a') else 'D'
                                                new_user = User.objects.create_alumniuserwithoutimage(email=row['email'], first_name=row['first_name'], last_name=row['last_name'], phone1=row['phone1'], password="Amahoro@1")
                                                family = Family.objects.get(id=row['familyid'])
                                                combination = Combination.objects.get(id=row['combinationid'])
                                                alumni_info = Alumni.objects.create(
                                                    user=new_user,
                                                    marital_status=row['marital_status'],
                                                    date_of_birth=row['date_of_birth'],
                                                    gender=row['gender'],
                                                    family=family,
                                                    combination=combination,
                                                    kids=True if row['kids'].lower().startswith('y') else False,
                                                    father="NN",
                                                    mother="NN",
                                                    did_you_born_in_rwanda=True if row['did_you_born_in_rwanda'].lower().startswith('y') else False,
                                                    place_of_birth_district_or_country=row['place_of_birth_district_or_country'],
                                                    place_of_birth_sector_or_city=row['place_of_birth_sector_or_city'],
                                                    currresidence_in_rwanda=True if row['currresidence_in_rwanda'].lower().startswith('y') else False,
                                                    currresidence_district_or_country=row['currresidence_district_or_country'],
                                                    currresidence_sector_or_city=row['currresidence_sector_or_city'],
                                                    s4marks=row['s4marks'], s5marks=row['s5marks'], s6marks=row['s6marks'], ne=row['ne'], maxforne=row['maxforne'],
                                                    decision='P' if row['decision'].lower().startswith('p') else 'F',
                                                    life_status=life)

                                                for ep_id in row['epsid']:
                                                    try:
                                                        ep = Ep.objects.get(id=ep_id)
                                                        alumni_info.eps.add(ep)
                                                    except Ep.DoesNotExist:
                                                        raise NotFound()

                                                if row['life_status'].lower().startswith('d'):
                                                    Studie.objects.create(alumn=alumni_info, level='D', degree="Deceased", university="Deceased", country="Deceased", scholarship='D', status="De", scholarship_details="Deceased")
                                                else:
                                                    if not isinstance(row['further_study_data']['study_level'], list):
                                                        # Handle the case where it's not a list
                                                        print(row['further_study_data']['study_level'])
                                                        Studie.objects.create(alumn=alumni_info, level='N', degree='NoInfo', university='NoInfo', country='NoInfo', scholarship='N', status="N", scholarship_details='NoInfo')
                                                    else:
                                                        st = pd.DataFrame(row['further_study_data'])
                                                        for index, row in st.iterrows():
                                                            Studie.objects.create(alumn=alumni_info, level=row['study_level'], degree=row['course_name'], university=row['university'], country=row['in_which_country'], scholarship=row['scholar'], status=row['study_status'], scholarship_details=row['scholar_details'])

                                                if life=="D":
                                                    Employment.objects.create(alumn=alumni_info, title="Deceased", status="D", description="Deceased", company="Deceased", career="Deceased")
                                                else:
                                                    if not isinstance(employdata['job_status'], list):
                                                        # Handle the case where it's not a list
                                                        Employment.objects.create(alumn=alumni_info, title='NoInfo', status="N", description='NoInfo', company='NoInfo', career='NoInfo')
                                                    else:
                                                        emp = pd.DataFrame(employdata)
                                                        for index, row in emp.iterrows():
                                                            Employment.objects.create(alumn=alumni_info, title=row['job_title'], status=row['job_status'], description='NN', company=row['company'], career=row['career'],end_date=row['current_old'])

                                               
                                            
                                else:
                                    data['error']="personal_profile sheet is empty!"
                                    print('personal_profile sheet is empty!')
                                    return Response(data)
                        else:
                            data['error']="Columns are not the same as the one expected"
                            print("Columns are not the same as the one expected")
                            return Response(data)

                    
                    else:
                        data["error"]="The sheet 'personal_profile' does not exist in the excel file."
                        print("The sheet 'personal_profile' does not exist in the excel file.")
                        return Response(data)
                    
                    
                else:
                    data["error"]="please upload using the real template"
                    print("please upload using the real template")
                    return Response(data)
                    

                return Response(data)
            else:
                data={"error":"The uploaded file is not an Excel file."}
                print("The uploaded file is not an Excel file.")
                return Response(data)

        return Response({'error': 'No file provided'})
    

class Generate_reports(APIView):
    permission_classes = [IsAuthenticated, ]
    def post(self, request):
        attrs=request.data
        # Initialize an empty list to store selected attributes
        selected_attributes = []

        # Iterate over attrs and add the attribute to selected_attributes if its corresponding value is True
        for attr, include in attrs.items():
            if include:
                selected_attributes.append(attr)
                
        if(len(selected_attributes)==0):
            return Response({"data":"No data"})

        # Construct the SELECT part of the SQL query
        select_query = ", ".join(selected_attributes)

        # Your base SQL query
        base_query = """
        SELECT
    {select_query}
FROM
    (SELECT
        api_user.email,
        api_user.first_name,
        api_user.last_name,
        api_user.phone1,
        api_user.image_url,
        userprofile_alumni.marital_status,
        userprofile_alumni.gender,
        userprofile_alumni.kids,
        userprofile_alumni.father,
        userprofile_alumni.mother,
        userprofile_alumni.s4marks,
        userprofile_alumni.s5marks,
        userprofile_alumni.s6marks,
        userprofile_alumni.ne,
        userprofile_alumni.maxforne,
        CASE 
            WHEN  userprofile_alumni.decision = 'P' THEN 'Pass'
            WHEN  userprofile_alumni.decision = 'F' THEN 'Fail'
            ELSE 'Unknown'
        END AS decision,
        CASE 
            WHEN userprofile_alumni.life_status = 'A' THEN 'Alive'
            WHEN userprofile_alumni.life_status = 'D' THEN 'Died'
            ELSE 'Unknown'
        END AS life_status,
        userprofile_alumni.family_id,
        userprofile_alumni.currresidence_district_or_country,
        userprofile_alumni.currresidence_in_rwanda,
        userprofile_alumni.currresidence_sector_or_city,
        userprofile_alumni.date_of_birth,
        userprofile_alumni.did_you_born_in_rwanda,
        userprofile_alumni.place_of_birth_district_or_country,
        userprofile_alumni.place_of_birth_sector_or_city,
        userprofile_combination.combination_name,
        userprofile_ep.title as ep_title, 
        CASE 
            WHEN userprofile_ep.type = 'A' THEN 'Arts'
            WHEN userprofile_ep.type = 'SC' THEN 'Sciences'
            WHEN userprofile_ep.type = 'S' THEN 'Sports'
            WHEN userprofile_ep.type = 'C' THEN 'Clubs'
            WHEN userprofile_ep.type = 'P' THEN 'Professional'
            ELSE 'Unknown'
        END AS ep_type,
        userprofile_family.id as fmly_id, 
        userprofile_family.family_name, 
        userprofile_family.family_number, 
        userprofile_family.family_mother, 
        userprofile_family.family_mother_tel, 
        userprofile_family.grade_id,
        userprofile_grade.id as gde_id, 
        userprofile_grade.grade_name, 
        userprofile_grade.start_academic_year, 
        userprofile_grade.end_academic_year,
        userprofile_employment.id as emp_id, 
        userprofile_employment.title as emp_title,
        CASE 
            WHEN userprofile_employment.status = 'F' THEN 'Full Time'
            WHEN userprofile_ep.type = 'P' THEN 'Part Time'
            WHEN userprofile_ep.type = 'S' THEN 'Self Employed'
            WHEN userprofile_ep.type = 'I' THEN 'Intern'
            WHEN userprofile_ep.type = 'U' THEN 'Unemployed'
            WHEN userprofile_ep.type = 'D' THEN 'Deceased'
            ELSE 'Unknown'
        END AS emp_status, 
        userprofile_employment.career, 
        userprofile_employment.description, 
        userprofile_employment.company, 
        userprofile_employment.start_date, 
        userprofile_employment.end_date, 
        userprofile_employment.alumn_id,
        userprofile_studie.id as st_id, 
        CASE 
            WHEN userprofile_studie.level = 'C' THEN 'Certificate'
            WHEN userprofile_studie.level = 'A1' THEN 'Advanced diploma'
            WHEN userprofile_studie.level = 'A0' THEN 'Bachelors'
            WHEN userprofile_studie.level = 'M' THEN 'Masters'
            WHEN userprofile_studie.level = 'NMS' THEN 'No More Study'
            WHEN userprofile_studie.level = 'PHD' THEN 'PHD'
            WHEN userprofile_studie.level = 'D' THEN 'Deceased'
            ELSE 'Unknown'
        END AS st_level,
        userprofile_studie.degree, 
        userprofile_studie.university, 
        userprofile_studie.country,
        CASE 
            WHEN userprofile_studie.scholarship = 'F' THEN 'Full'
            WHEN userprofile_studie.scholarship = 'P' THEN 'Partial'
            WHEN userprofile_studie.scholarship  = 'NS' THEN 'No Scholarship'
            WHEN userprofile_studie.scholarship  = 'NMS' THEN 'No More Study'
            WHEN userprofile_studie.scholarship = 'D' THEN 'Deceased'
            ELSE 'Unknown'
        END AS scholarship, 
        userprofile_studie.scholarship_details, 
        CASE 
            WHEN userprofile_studie.status = 'D' THEN 'Dropped_Out'
            WHEN userprofile_studie.status = 'D' THEN 'Suspended'
            WHEN userprofile_studie.status  = 'O' THEN 'On_going'
            WHEN userprofile_studie.status  = 'C' THEN 'Graduated'
            WHEN userprofile_studie.status  = 'NMS' THEN 'No Further Study'
            WHEN userprofile_studie.status = 'D' THEN 'Deceased'
            ELSE 'Unknown'
        END AS st_status
    FROM
        api_user
    LEFT JOIN
        userprofile_alumni ON api_user.id = userprofile_alumni.user_id
    LEFT JOIN
        userprofile_combination ON userprofile_alumni.combination_id=userprofile_combination.id
    LEFT JOIN
        userprofile_alumni_eps ON userprofile_alumni.id = userprofile_alumni_eps.alumni_id
    LEFT JOIN
        userprofile_ep ON userprofile_alumni_eps.ep_id=userprofile_ep.id
    LEFT JOIN
        userprofile_family ON userprofile_alumni.family_id=userprofile_family.id
    LEFT JOIN
        userprofile_grade ON userprofile_family.grade_id= userprofile_grade.id
    LEFT JOIN
        userprofile_employment ON userprofile_alumni.id=userprofile_employment.alumn_id
    LEFT JOIN
        userprofile_studie ON userprofile_alumni.id=userprofile_studie.alumn_id 
    WHERE
        api_user.is_alumni = TRUE) AS alumni_data;
        """.format(select_query=select_query)

        # Execute the SQL query
        with connection.cursor() as cursor:
            cursor.execute(base_query)
            rows = cursor.fetchall()
            unique_rows = []
            seen = set()  # Set to keep track of seen tuples

            for row in rows:
                # Convert tuple to string for hashing
                row_str = str(row)
                if row_str not in seen:
                    unique_rows.append(row)
                    seen.add(row_str)
            #print(unique_rows)
            # Create a new Workbook
            wb = Workbook()

            # Get the active worksheet
            ws = wb.active
            ws.append(selected_attributes)
            # Add rows
            alumni_data_name = 'alumni_data'
            ws.title = alumni_data_name
            for row_data in unique_rows:
                ws.append(row_data)

            # Save the workbook
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'
            wb.save(response)

            return response
        
        
        
#Library Management System

#TeacherOrLibrarian Reg. View
#useful link http://127.0.0.1:8000/api/bulkeducator/?id=129&?email=fifi@asyv.org
class TeacherOrLibrarianRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]
    def post(self, request):
        serializer = TeacherOrLibrarianRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        try:
            user_id = request.query_params.get('id')
            user_email = request.query_params.get('email')
            users_query = ("SELECT api_user.id as id, api_user.email as email, "
                           "api_user.phone1 as phone1, api_user.first_name as first_name, "
                           "api_user.last_name as last_name, api_user.image_url, "
                           "api_user.is_teacher, api_user.is_librarian "
                           "FROM api_user WHERE ")
            if user_id and user_email:
                users_query += f"api_user.id = {user_id} AND api_user.email = '{user_email}'"
            elif user_id:
                users_query += f"api_user.id = {user_id}"
            elif user_email:
                users_query += f"api_user.email = '{user_email}'"
            else:
                users_query += "api_user.is_teacher OR api_user.is_librarian"
                
            users = User.objects.raw(users_query)
            serializer = TeacherAndLibrarianSerializer(users, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
class TeacherOrLibrarianEditView(APIView):
    permission_classes = [IsAuthenticated, ]
    def put(self, request, *args, **kwargs):
        try:
            user_id = kwargs.get('id')
            user = User.objects.get(id=user_id)
            serializer = TeacherOrLibrarianRegistrationSerializer(user, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except User.DoesNotExist:
            return Response({'error': 'User does not exist'}, status=status.HTTP_404_NOT_FOUND)
        
class StudentRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]
    def post(self, request):
        serializer = StudentRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request):
        try:
            # Checking for the parameters from the URL
            if request.query_params:
                stud = Student.objects.filter(**request.query_params.dict())
                serializer = StudentSerializer(stud, many=True)
                return Response(serializer.data)
            else:
                stud = Student.objects.all()
                serializer = StudentSerializer(stud, many=True)
                return Response(serializer.data)
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
class StudentRegistrationUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def put(self, request, pk, format=None):
        try:
            user = User.objects.get(pk=pk)  # Provide pk as a keyword argument
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = StudentRegistrationSerializer(user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
# Author data view

class AuthorRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = AuthorSerializer(data=request.data)
        # validating for already existing data
        if Author.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                autho = Author.objects.filter(**request.query_params.dict())
            else:
                autho = Author.objects.all()

            # if there is something in items else raise error
            if autho:
                serializer = AuthorSerializer(autho, many=True)
                return Response(serializer.data)
            else:
                return Response([])
            
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_Author(request, pk):
    autho = Author.objects.get(pk=pk)
    data = AuthorSerializer(instance=autho, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_author(request, pk):
    autho = get_object_or_404(Author, pk=pk)
    autho.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


# end

# Category data view

class CategoryRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = CategorySerializer(data=request.data)
        # validating for already existing data
        if Category.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                cat = Category.objects.filter(**request.query_params.dict())
            else:
                cat = Category.objects.all()

            # if there is something in items else raise error
            if cat:
                serializer = CategorySerializer(cat, many=True)
                return Response(serializer.data)
            else:
                return Response([])
            
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_Category(request, pk):
    cat = Category.objects.get(pk=pk)
    data = CategorySerializer(instance=cat, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_category(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    cat.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


# end

# Book data view

class BookRegistrationView(APIView):
    permission_classes = [IsAuthenticated, ]

    def post(self, request):
        serializer = BookSerializer(data=request.data)
        # validating for already existing data
        if Book.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # checking for the parameters from the URL
            if request.query_params:
                book = Book.objects.filter(**request.query_params.dict())
            else:
                book = Book.objects.all()

            # if there is something in items else raise error
            if book:
                serializer = DisplayBookSerializer(book, many=True)
                return Response(serializer.data)
            else:
                return Response([])
            
        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
#@permission_classes([IsAuthenticated])
def update_Book(request, pk):
    book = Book.objects.get(pk=pk)
    data = BookSerializer(instance=book, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
#@permission_classes([IsAuthenticated])
def delete_book(request, pk):
    book = get_object_or_404(Book, pk=pk)
    book.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


# end

# Issue_Book data view

class Issue_BookRegistrationView(APIView):
    #permission_classes = [IsAuthenticated, ]
    pagination_class = CustomPagination  # Use custom pagination class

    def post(self, request):
        serializer = Issue_BookSerializer(data=request.data)
        # validating for already existing data
        if Issue_Book.objects.filter(**request.data).exists():
            raise serializers.ValidationError('This data already exists')

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        try:
            # Get the page number and page size from query parameters
            page_number = request.query_params.get('page', 1)
            page_size = request.query_params.get('page_size', 20)

            # Extract pagination parameters from the request query params
            query_params = request.query_params.copy()
            page_number = query_params.pop('page', 1)
            page_size = query_params.pop('page_size', 20)

            # Initialize filter conditions
            filter_conditions = {}

            # Extract special filter condition for studentid
            studentid = query_params.pop('studentid', None)
            if studentid:
                # Retrieve the Student instance based on the studentid
                student = Student.objects.get(studentid=studentid)
                # Access the associated User instance
                user = student.user
                # Add filter condition for borrower
                filter_conditions['borrower'] = user
            
            # Add the condition for 'returnedate' to filter for "Not yet Returned"
            filter_conditions['returndate'] = "Not yet Returned"

            # Apply remaining filter conditions
            for key, value in query_params.items():
                filter_conditions[key] = value

            # Filtering issue objects based on query params
            issue = Issue_Book.objects.filter(**filter_conditions)

            # Pagination
            paginator = self.pagination_class()
            paginated_issue = paginator.paginate_queryset(issue, request)

            # Serializing paginated data
            serializer = DisplayIssue_BookSerializer(paginated_issue, many=True)
            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_Issue_Book(request, pk):
    try:
        issue = Issue_Book.objects.get(pk=pk)
    except Issue_Book.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    data = {'returndate': request.data.get('returndate')}
    serializer = Issue_BookSerializer(instance=issue, data=data, partial=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_Issue_Book(request, pk):
    issue = get_object_or_404(Issue_Book, pk=pk)
    issue.delete()
    return Response(status=status.HTTP_202_ACCEPTED)


# end

#Check student

        
class CheckStudentView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get(self, request):
        try:
            user_email = request.query_params.get('email')
            if user_email:
                users = User.objects.filter(email=user_email, student__isnull=False).select_related('student__family__grade', 'student__combination')
            else:
                users = User.objects.filter(student__isnull=False).select_related('student__family__grade', 'student__combination')
                
            serializer = CheckStudentSerializer(users, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class ChangeStudentPasswordView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def post(self, request):
        serializer = ChangeStudentPasswordSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            password = serializer.validated_data['password']
            user = User.objects.filter(email=email).first()
            if user:
                user.set_password(password)
                user.save()
                return Response({"message": "Password changed successfully"}, status=status.HTTP_200_OK)
            else:
                return Response({"message": "User not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class UserList(generics.ListAPIView):
    # permission_classes = [IsAuthenticated, ]
    serializer_class = UsersSerializer

    def get_queryset(self):
        student_id = self.request.query_params.get('student_info__studentid', None)
        if student_id is not None:
            queryset = User.objects.filter(student__studentid=student_id).prefetch_related('borrow', 'student').all()
            return queryset
        else:
            raise Http404("Please provide a valid student ID.")

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.serializer_class(queryset, many=True)
        return Response(serializer.data)
    
class AutoDataExcelUploadAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    parser_classes = (MultiPartParser,)

    def post(self, request, format=None):
        file = request.FILES.get('file')
        if file:
            if is_excel_file(file):
                print("The uploaded file is an Excel file.")
                
                workbook = openpyxl.load_workbook(file)
                sheet_names = workbook.sheetnames
                data = {}
                expected_sheet_names=['author','category','book']
                if set(sheet_names) ==set(expected_sheet_names):
                    for sheet_name in sheet_names:
                        sheet = workbook[sheet_name]
                        rows = []

                        for row in sheet.iter_rows(values_only=True):
                            rows.append(row)

                        data[sheet_name] = rows
                    
                    #...........validate author data.........................
                    
                    if 'author' in data:
                        # Extract headers and data
                        author_headers = data['author'][0]
                        if set(data['author'][0]) == set(('id', 'author_name')):
                            author_rows=data['author'][1:]
                            if(len(author_rows)>0):
                                # Create a DataFrame
                                author= pd.DataFrame(author_rows, columns=author_headers)
                                
                                # Check for null values in the entire DataFrame
                                # Sum the null values in each column
                                null_countsauthor = author.isnull().sum()
                                if(null_countsauthor>0).all():
                                    print("\nNull counts in each column in profile:")
                                    print(null_countsauthor)
                                    data["error"]="there are some empty cells in author"
                                    return Response(data)
                                else:
                                    #Data validation on Category
                                    # Extract headers and data
                                    category_headers = data['category'][0]
                                    if(set(category_headers)==set(('id', 'category_name'))):
                                        category_rows=data['category'][1:]
                                        if(len(category_rows)>0):
                                            # Create a DataFrame
                                            category= pd.DataFrame(category_rows, columns=category_headers)
                                            
                                            # Check for null values in the entire DataFrame
                                            # Sum the null values in each column
                                            null_countscategory = category.isnull().sum()
                                            if(null_countscategory>0).all():
                                                print("\nNull counts in each column in category:")
                                                print(null_countscategory)
                                                data["error"]="there are some empty cells in category"
                                                return Response(data)
                                            else:     
                                                #Data validation on Book
                                                # Extract headers and data
                                                book_headers = data['book'][0]
                                                if(set(book_headers)==set(('id', 'book_name', 'isbnumber', 'number_of_books', 'category', 'author'))):
                                                    book_rows=data['book'][1:]
                                                    if(len(book_rows)>0):
                                                        # Create a DataFrame
                                                        book= pd.DataFrame(book_rows, columns=book_headers)
                                                        
                                                        null_countsbook = book.isnull().sum()
                                                        if(null_countsbook>0).all():
                                                            print("\nNull counts in each column in book:")
                                                            print(null_countsbook)
                                                            data["error"]="there are some empty cells in book"
                                                            return Response(data)
                                                        else:
                                       
                                                            for index, row in author.iterrows():
                                                                #print(f"Index: {index}, id: {row['id']} Author Name: {row['author_name']}")
                                                                Author.objects.create(author_name=row['author_name'],author_id=row['id'])
                                                            for index, row in category.iterrows():
                                                                #print(f"Index: {index}, id: {row['id']} Author Name: {row['author_name']}")
                                                                Category.objects.create(category_name=row['category_name'],category_id=row['id'])
                                                                
                                                            for index, row in book.iterrows():
                                                                #print(f"Index: {index}, id: {row['id']} Author Name: {row['author_name']}")
                                                                try:
                                                                    # Attempt to retrieve the category instance
                                                                    category = Category.objects.get(category_id=row['category'])
                                                                except Category.DoesNotExist:
                                                                    # Handle the case where the author doesn't exist
                                                                    print("Category with ID {} does not exist".format(row['id']))
                                                                    continue;
                                                                try:
                                                                    # Attempt to retrieve the author instance
                                                                    author = Author.objects.get(author_id=row['author'])
                                                                except Author.DoesNotExist:
                                                                    # Handle the case where the author doesn't exist
                                                                    print("Author with ID {} does not exist".format(row['id']))
                                                                    continue;
                                                                Book.objects.create(book_name=row['book_name'],isbnumber=row['isbnumber'],number_of_books=row['number_of_books'],category=category,author=author)
                                          
                                                    else:
                                                        data['error']="book sheet is empty!"
                                                        print('book sheet is empty!')
                                                        return Response(data)
                     
                                    
                                    
                     
            else:
                data={"error":"The uploaded file is not an Excel file."}
                print("The uploaded file is not an Excel file.")
                return Response(data)
        return Response({"error": "Data uploaded successfully"})
    
class AutoStudentDataExcelUploadAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    parser_classes = (MultiPartParser,)

    def post(self, request, format=None):
        file = request.FILES.get('file')
        data = {}

        if not file:
            data["error"] = "The uploaded file is missing."
            return Response(data)

        if not is_excel_file(file):
            data["error"] = "The uploaded file is not an Excel file."
            return Response(data)

        try:
            workbook = openpyxl.load_workbook(file)
            sheet_names = workbook.sheetnames
            expected_sheet_names = ['students']

            if set(sheet_names) != set(expected_sheet_names):
                data["error"] = "The uploaded file contains different sheets."
                return Response(data)

            sheet = workbook['students']
            df_students = pd.DataFrame(sheet.iter_rows(min_row=2, values_only=True), columns=[cell for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0])

            if not df_students.empty:
                if set(df_students.columns) != set(["email", "first_name", "last_name", "phone1", "password","gender", "studentid", "family", "combination"]):
                    data["error"] = "Students sheets have different headers."
                    data["come"] = df_students.columns
                    data["going"] = set(["email", "first_name", "last_name", "phone1", "password","gender", "studentid", "family", "combination"])
                    return Response(data)

                families = df_students['family'].unique()
                combinations = df_students['combination'].unique()
                missing_families = []
                missing_combinations = []

                for family_name in families:
                    if not Family.objects.filter(family_name=family_name).exists():
                        missing_families.append(family_name)

                for combination_name in combinations:
                    if not Combination.objects.filter(combination_name=combination_name).exists():
                        missing_combinations.append(combination_name)

                if missing_families or missing_combinations:
                    data["error"] = {
                        "missing_families": missing_families,
                        "missing_combinations": missing_combinations
                    }
                    return Response(data)

                # Now create student instances since all families and combinations exist
                for index, row in df_students.iterrows():
                    family = Family.objects.get(family_name=row['family'])
                    combination = Combination.objects.get(combination_name=row['combination'])
                    user = User.objects.create_studentuserwithoutimage(row['email'],row['first_name'], row['last_name'], row['phone1'], row['password'], row['gender'])
                    Student.objects.create(user=user, family=family, combination=combination, studentid=row['studentid'])

            else:
                data["error"] = "Students sheet is empty!"
                return Response(data)

        except Exception as e:
            data["error"] = str(e)
            return Response(data)

        return Response({"msg": "Data uploaded successfully"})
    

class AutoIssueDataExcelUploadAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    parser_classes = (MultiPartParser,)

    def post(self, request, format=None):
        file = request.FILES.get('file')
        data = {}

        if not file:
            data["error"] = "The uploaded file is missing."
            return Response(data)

        if not is_excel_file(file):
            data["error"] = "The uploaded file is not an Excel file."
            return Response(data)

        try:
            workbook = openpyxl.load_workbook(file)
            sheet_names = workbook.sheetnames
            expected_sheet_names = ['issue']

            if set(sheet_names) != set(expected_sheet_names):
                data["error"] = "The uploaded file contains different sheets."
                return Response(data)

            sheet = workbook['issue']
            df_issue = pd.DataFrame(sheet.iter_rows(min_row=2, values_only=True), columns=[cell for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0])

            if not df_issue.empty:
                if set(df_issue.columns) != set(["book", 'borrower', 'library_number', 'issuedate', 'returndate']):
                    data["error"] = "Issue sheets have different headers."
                    return Response(data)

                books = df_issue['book'].unique()
                borrowers = df_issue['borrower'].unique()
                missing_books = []
                missing_borrowers = []

                for book_id in books:
                    if not Book.objects.filter(isbnumber=book_id).exists():
                        missing_books.append(book_id)

                for borrower in borrowers:
                    # Query the Student model based on the studentid
                    try:
                        student = Student.objects.get(studentid=borrower)
                        user_instance = student.user  # Access the User instance associated with the Student
                        # Check if user_instance is not None (i.e., it exists)
                        if user_instance is not None:
                            print(user_instance)
                        else:
                            # If user_instance is None, handle the case where the associated User doesn't exist
                            missing_borrowers.append(borrower)
                    except Student.DoesNotExist:
                        # Handle the case where no Student with the given studentid exists
                        missing_borrowers.append(borrower)

                if missing_books or missing_borrowers:
                    data["error"] = {
                        "missing_books": missing_books,
                        "missing_borrowers": missing_borrowers
                    }
                    return Response(data)

                # Now create Issue instances since all books and borrowers exist
                for index, row in df_issue.iterrows():
                    book = Book.objects.get(isbnumber=row['book'])
                    student = Student.objects.get(studentid=row['borrower'])
                    user_instance = student.user 
                    Issue_Book.objects.create(book=book, borrower=user_instance, library_number=row['library_number'], issuedate=row['issuedate'], returndate=row['returndate'])

            else:
                data["error"] = "Issue sheet is empty!"
                return Response(data)

        except Exception as e:
            data["error"] = str(e)
            return Response(data)

        return Response({"msg": "Data uploaded successfully"})
    
    
#report data from library database 
class IssuedBookDisplayAPIView(APIView):
    permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:
            # Get data
            sql_query1 = """
                SELECT grade_name, studentid, family_name, combination_name, first_name, last_name, 
                       email, book_name, isbnumber, category_name, author_name, library_number, 
                       issuedate, returndate, userprofile_issue_book.id AS id 
                FROM api_user 
                INNER JOIN userprofile_issue_book ON api_user.id = userprofile_issue_book.borrower_id 
                INNER JOIN userprofile_student ON api_user.id = userprofile_student.user_id 
                INNER JOIN userprofile_family ON userprofile_student.family_id = userprofile_family.id 
                INNER JOIN userprofile_grade ON userprofile_family.grade_id = userprofile_grade.id 
                INNER JOIN userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id 
                INNER JOIN userprofile_book ON userprofile_issue_book.book_id = userprofile_book.id 
                INNER JOIN userprofile_category ON userprofile_book.category_id = userprofile_category.id 
                INNER JOIN userprofile_author ON userprofile_book.author_id = userprofile_author.id 
                WHERE returndate = 'Not yet Returned' order by issuedate desc;
            """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query1)
                data1 = cursor.fetchall()

            data = []
            if data1 is not None:
                for i in data1:
                    data.append({
                        'grade_name': i[0],
                        'studentid': i[1],
                        'family_name': i[2],
                        'combination_name': i[3],
                        'first_name': i[4],
                        'last_name': i[5],
                        'email': i[6],
                        'book_name': i[7],
                        'isbnumber': i[8],
                        'category_name': i[9],
                        'author_name': i[10],
                        'library_number': i[11],
                        'issuedate': i[12],
                        'returndate': i[13],
                        'id': i[14]
                    })

            serializer = IssuedBookDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        
class StudentListDisplayAPIView(APIView):
    permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:
            # Get data
            sql_query1 = """
                select grade_name,studentid,family_name,combination_name,
                first_name,last_name,email,userprofile_student.id as id,userprofile_grade.id as grade_id,userprofile_combination.id as combination_id,userprofile_grade.end_academic_year as eay,gender from 
                api_user  inner join userprofile_student on 
                api_user.id=userprofile_student.user_id inner join 
                userprofile_family on userprofile_student.family_id=userprofile_family.id inner join 
                userprofile_grade on userprofile_family.grade_id=userprofile_grade.id inner join 
                userprofile_combination on userprofile_student.combination_id=userprofile_combination.id where api_user.current;
            """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query1)
                data1 = cursor.fetchall()

            data = []
            if data1 is not None:
                for i in data1:
                    data.append({
                        'grade_name': i[0],
                        'studentid': i[1],
                        'family_name': i[2],
                        'combination_name': i[3],
                        'first_name': i[4],
                        'last_name': i[5],
                        'email': i[6],
                        'id': i[7],
                        'grade_id': i[8],
                        'combination_id': i[9],
                        'eay': i[10],
                        'gender': i[11]
                    })

            serializer = StudentListDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        
class BookListDisplayAPIView(APIView):
    permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:
            # Get data
            sql_query1 = """
                SELECT book_name, isbnumber, category_name, author_name, number_of_books, 
                       userprofile_book.id AS id  FROM  userprofile_book 
                INNER JOIN userprofile_category ON userprofile_book.category_id = userprofile_category.id 
                INNER JOIN userprofile_author ON userprofile_book.author_id = userprofile_author.id  
                order by book_name asc;
            """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query1)
                data1 = cursor.fetchall()

            data = []
            if data1 is not None:
                for i in data1:
                    data.append({
                        'book_name': i[0],
                        'isbnumber': i[1],
                        'category_name': i[2],
                        'author_name': i[3],
                        'number_of_books': i[4],
                        'id': i[5]
                    })

            serializer = BookListDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        

class BookReportExportAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get_data_from_database(self):
        sql_query = """
            select
                ROW_NUMBER() OVER (ORDER BY category_name ASC) AS "Number",
                userprofile_book.book_name,
                userprofile_book.isbnumber,
                userprofile_category.category_name,
                userprofile_author.author_name,
                userprofile_book.number_of_books::int, 
                COUNT(userprofile_issue_book.id) AS issued_books,
                (userprofile_book.number_of_books::int - COUNT(userprofile_issue_book.id)) AS current_books
            FROM
                userprofile_book
            INNER JOIN
                userprofile_category ON userprofile_book.category_id = userprofile_category.id
            INNER JOIN
                userprofile_author ON userprofile_book.author_id = userprofile_author.id
            LEFT JOIN
                userprofile_issue_book ON userprofile_issue_book.book_id = userprofile_book.id AND userprofile_issue_book.returndate = 'Not yet Returned'
            GROUP BY
                userprofile_book.book_name,
                userprofile_book.isbnumber,
                userprofile_category.category_name,
                userprofile_author.author_name,
                userprofile_book.number_of_books::int -- Casting to integer
            ORDER BY
                category_name ASC;
        """
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            data = cursor.fetchall()

        # Calculate totals
        total_books = sum(row[5] for row in data)
        total_issued = sum(row[6] for row in data)
        total_current = sum(row[7] for row in data)

        # Append totals as a new row
        total_row = ["", "Total", "", "", "", total_books, total_issued, total_current]
        data.append(total_row)

        return data

    def generate_pdf(self, data):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="list_of_books.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        # Add title
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_paragraph = Paragraph("LFHS@ASYV Library List of Books", title_style)
        elements.append(title_paragraph)

        # Add data table
        table_style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        
        table_data = [['#','Book Name', 'ISBN Number', 'Category', 'Author', 'Number of Books', 'Issued Books','current_books']]
        table_data.extend(data)

        # Calculate maximum column widths based on available page width
        available_width = doc.width
        num_cols = len(table_data[0])
        max_col_width = available_width / num_cols

        # Add table content with wrapped paragraphs
        wrapped_table_data = []
        for row in table_data:
            wrapped_row = []
            for cell in row:
                cell_style = ParagraphStyle(name='WrapStyle', wordWrap='LTR')
                wrapped_cell = Paragraph(str(cell), cell_style)
                wrapped_row.append(wrapped_cell)
            wrapped_table_data.append(wrapped_row)

        table = Table(wrapped_table_data)
        table.setStyle(table_style)

        elements.append(table)
        doc.build(elements)
        return response

    def get(self, request, *args, **kwargs):
        try:
            data = self.get_data_from_database()
            if data:
                return self.generate_pdf(data)
            else:
                return Response({'error': 'No data found.'}, status=404)
        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)

class Issued_BookReportExportAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get_data_from_database(self):
        sql_query = """
            SELECT 
                grade_name,family_name,combination_name,studentid, last_name, first_name,book_name,isbnumber,
                category_name, author_name,issuedate,returndate 
            FROM 
                userprofile_grade 
            INNER JOIN 
                userprofile_family ON userprofile_grade.id = userprofile_family.grade_id 
            INNER JOIN 
                userprofile_student ON userprofile_family.id = userprofile_student.family_id 
            INNER JOIN 
                api_user ON userprofile_student.user_id = api_user.id 
            INNER JOIN 
                userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id 
            INNER JOIN 
                userprofile_issue_book ON userprofile_issue_book.borrower_id = api_user.id 
            INNER JOIN 
                userprofile_book ON userprofile_issue_book.book_id = userprofile_book.id 
            INNER JOIN 
                userprofile_category ON userprofile_book.category_id = userprofile_category.id 
            INNER JOIN 
                userprofile_author ON userprofile_book.author_id = userprofile_author.id 
            WHERE 
                returndate = 'Not yet Returned' 
            ORDER BY 
                grade_name ASC, 
                family_name ASC;
        """
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            data = cursor.fetchall()

        return data

    def generate_pdf(self, data):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="list_of_Issued_books.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        # Define a base style for titles
        base_title_style = getSampleStyleSheet()['Title']

        # Define different title styles with different font sizes
        title_style_small = ParagraphStyle(
            'TitleSmall',
            parent=base_title_style,
            fontSize=12  # Set the font size to 12
        )

        title_style_medium = ParagraphStyle(
            'TitleMedium',
            parent=base_title_style,
            fontSize=14  # Set the font size to 18
        )
        title_paragraph = Paragraph("LFHS@ASYV Library List of Issued Books", base_title_style)
        elements.append(title_paragraph)
        table_style = TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                  ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                  ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                  ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                  ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                  ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                  ('GRID', (0, 0), (-1, -1), 1, colors.black)])

        # Group data by grade_name
        grouped_data = {}
        for row in data:
            grade_name = row[0]  # Assuming grade_name is in the first position of each tuple
            family_name = row[1]  # Assuming family_name is in the second position of each tuple
            if grade_name not in grouped_data:
                grouped_data[grade_name] = {}
            if family_name not in grouped_data[grade_name]:
                grouped_data[grade_name][family_name] = []
            # Exclude grade_name and family_name when extending the list
            grouped_data[grade_name][family_name].append(tuple(row[2:]))  # Start from the third element
            
        # Log grouped_data
        
        available_width = doc.width
        # Create tables for each grade_name
        for grade_name, families in grouped_data.items():
            # Add grade_name as title
            grade_title = Paragraph(grade_name+" Grade", title_style_medium)
            elements.append(grade_title)

            # Create tables for each family_name
            for family_name, family_data in families.items():
                # Add family_name as title
                family_title = Paragraph("Books Issued in "+family_name+" Family from "+grade_name+" Grade", title_style_small)
                elements.append(family_title)

                # Prepare data for table
                table_data = [['#','Student ID', 'Name','Book Name', 'ISBN Number', 'Category', 'Author',
                               'Issue Date', 'Return Date']]
                for idx, item in enumerate(family_data, start=1):
                    #logging.debug("Item tuple: %s", item)  # Log the contents of item
                    table_data.append([
                        idx,
                        item[1],
                        item[2]+" "+item[3]+" ("+item[0]+")",
                        item[4],
                        item[5],
                        item[6],
                        item[7],
                        item[8],
                        item[9]
                    ])
                
                # Calculate maximum column widths based on available page width
                
                num_cols = len(table_data[0])
                max_col_width = available_width / num_cols

                # Add table content with wrapped paragraphs
                wrapped_table_data = []
                for row in table_data:
                    wrapped_row = []
                    for cell in row:
                        cell_style = ParagraphStyle(name='WrapStyle', wordWrap='LTR')
                        wrapped_cell = Paragraph(str(cell), cell_style)
                        wrapped_row.append(wrapped_cell)
                    wrapped_table_data.append(wrapped_row)

                table = Table(wrapped_table_data)
                table.setStyle(table_style)

                elements.append(table)
                

        # Build the PDF document
        doc.build(elements)

        return response

    def get(self, request, *args, **kwargs):
        try:
            data = self.get_data_from_database()
            if data:
                return self.generate_pdf(data)
            else:
                return Response({'error': 'No data found.'}, status=404)
        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        
class Overdue_BookReportExportAPIView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get_data_from_database(self):
        sql_query = """
           select 
                grade_name,
                family_name,
                combination_name,
                studentid,
                last_name,
                first_name,
                book_name,
                isbnumber,
                category_name,
                author_name,
                issuedate,
                returndate,
                (NOW()::date - issuedate::date) AS days_overdue
            FROM 
                userprofile_grade 
            INNER JOIN 
                userprofile_family ON userprofile_grade.id = userprofile_family.grade_id 
            INNER JOIN 
                userprofile_student ON userprofile_family.id = userprofile_student.family_id 
            INNER JOIN   
                api_user ON userprofile_student.user_id = api_user.id 
            INNER JOIN 
                userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id 
            INNER JOIN  
                userprofile_issue_book ON userprofile_issue_book.borrower_id = api_user.id 
            INNER JOIN  
                userprofile_book ON userprofile_issue_book.book_id = userprofile_book.id 
            INNER JOIN 
                userprofile_category ON userprofile_book.category_id = userprofile_category.id  
            INNER JOIN  
                userprofile_author ON userprofile_book.author_id = userprofile_author.id  
            WHERE  
                returndate = 'Not yet Returned' 
                AND (NOW()::date - issuedate::date) > 30  -- Comparing the difference in days to 30
            ORDER BY   
                grade_name ASC,  
                family_name ASC;
        """
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            data = cursor.fetchall()

        return data

    def generate_pdf(self, data):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="list_of_Issued_books.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        # Define a base style for titles
        base_title_style = getSampleStyleSheet()['Title']

        # Define different title styles with different font sizes
        title_style_small = ParagraphStyle(
            'TitleSmall',
            parent=base_title_style,
            fontSize=12  # Set the font size to 12
        )

        title_style_medium = ParagraphStyle(
            'TitleMedium',
            parent=base_title_style,
            fontSize=14  # Set the font size to 18
        )
        title_paragraph = Paragraph("LFHS@ASYV Library List of Issued Books", base_title_style)
        elements.append(title_paragraph)
        table_style = TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                  ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                  ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                  ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                  ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                  ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                  ('GRID', (0, 0), (-1, -1), 1, colors.black)])

        # Group data by grade_name
        grouped_data = {}
        for row in data:
            grade_name = row[0]  # Assuming grade_name is in the first position of each tuple
            family_name = row[1]  # Assuming family_name is in the second position of each tuple
            if grade_name not in grouped_data:
                grouped_data[grade_name] = {}
            if family_name not in grouped_data[grade_name]:
                grouped_data[grade_name][family_name] = []
            # Exclude grade_name and family_name when extending the list
            grouped_data[grade_name][family_name].append(tuple(row[2:]))  # Start from the third element
            
        # Log grouped_data
        
        available_width = doc.width
        # Create tables for each grade_name
        for grade_name, families in grouped_data.items():
            # Add grade_name as title
            grade_title = Paragraph(grade_name+" Grade", title_style_medium)
            elements.append(grade_title)

            # Create tables for each family_name
            for family_name, family_data in families.items():
                # Add family_name as title
                family_title = Paragraph("Books Issued in "+family_name+" Family from "+grade_name+" Grade", title_style_small)
                elements.append(family_title)

                # Prepare data for table
                table_data = [['#','Student ID', 'Name','Book Name', 'ISBN Number', 'Category', 'Author',
                               'Issue Date', 'days_overdue']]
                for idx, item in enumerate(family_data, start=1):
                    #logging.debug("Item tuple: %s", item)  # Log the contents of item
                    table_data.append([
                        idx,
                        item[1],
                        item[2]+" "+item[3]+" ("+item[0]+")",
                        item[4],
                        item[5],
                        item[6],
                        item[7],
                        item[8],
                        item[10]
                    ])
                
                # Calculate maximum column widths based on available page width
                
                num_cols = len(table_data[0])
                max_col_width = available_width / num_cols

                # Add table content with wrapped paragraphs
                wrapped_table_data = []
                for row in table_data:
                    wrapped_row = []
                    for cell in row:
                        cell_style = ParagraphStyle(name='WrapStyle', wordWrap='LTR')
                        wrapped_cell = Paragraph(str(cell), cell_style)
                        wrapped_row.append(wrapped_cell)
                    wrapped_table_data.append(wrapped_row)

                table = Table(wrapped_table_data)
                table.setStyle(table_style)

                elements.append(table)
                

        # Build the PDF document
        doc.build(elements)

        return response

    def get(self, request, *args, **kwargs):
        try:
            data = self.get_data_from_database()
            if data:
                return self.generate_pdf(data)
            else:
                return Response({'error': 'No data found.'}, status=404)
        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        
class GeneralReportDisplayAPIView(APIView):
    permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:
            # Combined SQL Query
            sql_query = """
                with book_data as (
                    select 
                        count(userprofile_book.id) as nbook_types, 
                        sum(cast(userprofile_book.number_of_books as integer)) as nbooks 
                    from 
                        userprofile_book
                ),
                student_data as (
                    select 
                        count(api_user.id) as nstudents 
                    from 
                        api_user 
                    where 
                        api_user.is_student and api_user.current
                ),
                issue_data as (
                    select 
                        count(userprofile_issue_book.id) as nissued_books,
                        sum(case when (now()::date - issuedate::date) > 30 then 1 else 0 end) as noverdue_books 
                    from 
                        userprofile_issue_book 
                    where 
                        returndate = 'Not yet Returned'
                )
                select 
                    bd.nbook_types,
                    bd.nbooks,
                    sd.nstudents,
                    id.nissued_books,
                    id.noverdue_books
                from 
                    book_data bd,
                    student_data sd,
                    issue_data id;
            """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                data1 = cursor.fetchall()

            data = []
            if data1:
                data.append({
                    "nbook_types": data1[0][0],
                    "nbooks": data1[0][1],
                    "nstudents": data1[0][2],
                    "nissued_books": data1[0][3],
                    "noverdue_books": data1[0][4]
                })

            serializer = GeneralReportDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        
class StudentsReportExportAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get_data_from_database(self):
        sql_query = """
           SELECT 
                ROW_NUMBER() OVER (ORDER BY userprofile_grade.grade_name, userprofile_combination.combination_name) AS number,
                api_user.email,
                api_user.last_name,
                api_user.first_name,
                userprofile_student.studentid,
                userprofile_grade.grade_name,
                userprofile_family.family_name,
                userprofile_combination.combination_name
            FROM 
                api_user
            INNER JOIN 
                userprofile_student ON api_user.id = userprofile_student.user_id
            INNER JOIN 
                userprofile_family ON userprofile_family.id = userprofile_student.family_id
            INNER JOIN 
                userprofile_combination ON userprofile_combination.id = userprofile_student.combination_id
            INNER JOIN 
                userprofile_grade ON userprofile_grade.id = userprofile_family.grade_id
            WHERE 
                api_user.is_student
            ORDER BY 
                userprofile_grade.grade_name, 
                userprofile_combination.combination_name;

        """
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            data = cursor.fetchall()

        return data

    def generate_excel(self, data):
        # Create a new Workbook
        wb = Workbook()

        # Get the active worksheet
        ws = wb.active
        ws.append(["No","Email", "Last Name", "First Name", "Reg.No", "Grade", "Family", "Class"])
        # Add rows
        students_data_name = 'students_data'
        ws.title = students_data_name
        #logging.debug(data)
        
        for row_data in data:
            ws.append(row_data)

        # Save the workbook
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=LFHS_students_data.xlsx'
        wb.save(response)
        return response

    def get(self, request, *args, **kwargs):
        try:
            data = self.get_data_from_database()
            if data:
                return self.generate_excel(data)
            else:
                return Response({'error': 'No data found.'}, status=404)
        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)

#example of url: api/mostborrower/ or api/mostborrower/?start_date=2023-09-25T00:00:00&end_date=2024-05-31T23:59:59
class MostBorrowerDisplayAPIView(APIView):
    # permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:
            # Get query parameters for start_date and end_date
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')

            if start_date and end_date:
                # Ensure start_date and end_date are in the correct format
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%S')
                    end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S')
                except ValueError:
                    return Response({'error': 'Invalid date-time format. Use YYYY-MM-DDTHH:MM:SS.'}, status=400)
                
                # SQL query for the specified date range
                sql_query = f"""
                    SELECT 
                            api_user.last_name, 
                        api_user.first_name, 
                        userprofile_grade.grade_name, 
                        userprofile_family.family_name, 
                        userprofile_combination.combination_name, 
                        COUNT(userprofile_issue_book.id) AS issue_count
                        FROM 
                            api_user
                        INNER JOIN 
                            userprofile_issue_book 
                        ON 
                            api_user.id = userprofile_issue_book.borrower_id
                        INNER JOIN 
                            userprofile_student 
                        ON 
                            api_user.id = userprofile_student.user_id
                        INNER JOIN 
                            userprofile_family 
                        ON 
                            userprofile_student.family_id = userprofile_family.id
                        INNER JOIN 
                            userprofile_grade 
                        ON 
                            userprofile_family.grade_id = userprofile_grade.id
                        INNER JOIN 
                            userprofile_combination 
                        ON 
                            userprofile_student.combination_id = userprofile_combination.id
                        WHERE 
                            api_user.is_student = true
                            AND userprofile_issue_book.issuedate >= '{start_date}'
                            AND userprofile_issue_book.issuedate <= '{end_date}'
                        GROUP BY 
                        api_user.last_name, 
                        api_user.first_name, 
                        userprofile_grade.grade_name, 
                        userprofile_family.family_name, 
                        userprofile_combination.combination_name
                        ORDER BY 
                            issue_count DESC
                        LIMIT 5
                """
            else:
                # SQL query for the current month
                sql_query = """
                    SELECT 
                            api_user.last_name, 
                        api_user.first_name, 
                        userprofile_grade.grade_name, 
                        userprofile_family.family_name, 
                        userprofile_combination.combination_name, 
                        COUNT(userprofile_issue_book.id) AS issue_count
                        FROM 
                            api_user
                        INNER JOIN 
                            userprofile_issue_book 
                        ON 
                            api_user.id = userprofile_issue_book.borrower_id
                        INNER JOIN 
                            userprofile_student 
                        ON 
                            api_user.id = userprofile_student.user_id
                        INNER JOIN 
                            userprofile_family 
                        ON 
                            userprofile_student.family_id = userprofile_family.id
                        INNER JOIN 
                            userprofile_grade 
                        ON 
                            userprofile_family.grade_id = userprofile_grade.id
                        INNER JOIN 
                            userprofile_combination 
                        ON 
                            userprofile_student.combination_id = userprofile_combination.id
                        WHERE 
                            api_user.is_student = true
                            AND DATE_TRUNC('month', CAST(userprofile_issue_book.issuedate AS DATE)) = DATE_TRUNC('month', CURRENT_DATE)
                        GROUP BY 
                        api_user.last_name, 
                        api_user.first_name, 
                        userprofile_grade.grade_name, 
                        userprofile_family.family_name, 
                        userprofile_combination.combination_name
                        ORDER BY 
                            issue_count DESC
                        LIMIT 5
                """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                data1 = cursor.fetchall()

            data = []
            if data1 is not None:
                for i in data1:
                    data.append({
                        'last_name': i[0],
                        'first_name': i[1],
                        'grade_name': i[2],
                        'family_name': i[3],
                        'combination_name': i[4],
                        'issue_count': i[5]
                    })

            serializer = MostBorrowerDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)

class BorrowerByGradeDisplayAPIView(APIView):
    # permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:
            # Get query parameters for start_date and end_date
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')

            if start_date and end_date:
                # Ensure start_date and end_date are in the correct format
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%S')
                    end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S')
                except ValueError:
                    return Response({'error': 'Invalid date-time format. Use YYYY-MM-DDTHH:MM:SS.'}, status=400)
                
                # SQL query for the specified date range
                sql_query = f"""
                    SELECT 
                        userprofile_grade.grade_name,
                        userprofile_family.family_name,
                        userprofile_combination.combination_name,
                        COUNT(DISTINCT(userprofile_issue_book.borrower_id)) AS borrows,
                        COUNT(DISTINCT(api_user.id)) AS students
                    FROM 
                        api_user
                    INNER JOIN 
                        userprofile_student ON api_user.id = userprofile_student.user_id
                    INNER JOIN 
                        userprofile_family ON userprofile_student.family_id = userprofile_family.id
                    INNER JOIN  
                        userprofile_grade ON userprofile_family.grade_id = userprofile_grade.id
                    INNER JOIN 
                        userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id
                    LEFT JOIN 
                        userprofile_issue_book ON api_user.id = userprofile_issue_book.borrower_id 
                        AND userprofile_issue_book.issuedate >= '{start_date}'
                        AND userprofile_issue_book.issuedate <= '{end_date}'
                    WHERE 
                        api_user.is_student = true
                    GROUP BY 
                        userprofile_grade.grade_name,
                        userprofile_family.family_name,
                        userprofile_combination.combination_name
                    ORDER BY 
                        userprofile_grade.grade_name ASC,
                        userprofile_family.family_name ASC,
                        userprofile_combination.combination_name ASC,  
                        borrows DESC;
                """
            else:
                # SQL query for the current month
                sql_query = """
                    SELECT 
                        userprofile_grade.grade_name,
                        userprofile_family.family_name,
                        userprofile_combination.combination_name,
                        COUNT(DISTINCT(userprofile_issue_book.borrower_id)) AS borrows,
                        COUNT(DISTINCT(api_user.id)) AS students
                    FROM 
                        api_user
                    INNER JOIN 
                        userprofile_student ON api_user.id = userprofile_student.user_id
                    INNER JOIN 
                        userprofile_family ON userprofile_student.family_id = userprofile_family.id
                    INNER JOIN  
                        userprofile_grade ON userprofile_family.grade_id = userprofile_grade.id
                    INNER JOIN 
                        userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id
                    LEFT JOIN 
                        userprofile_issue_book ON api_user.id = userprofile_issue_book.borrower_id 
                        AND DATE_TRUNC('month', CAST(userprofile_issue_book.issuedate AS DATE)) = DATE_TRUNC('month', CURRENT_DATE)
                    WHERE 
                        api_user.is_student = true
                    GROUP BY 
                        userprofile_grade.grade_name,
                        userprofile_family.family_name,
                        userprofile_combination.combination_name
                    ORDER BY 
                        userprofile_grade.grade_name ASC,
                        userprofile_family.family_name ASC,
                        userprofile_combination.combination_name ASC,  
                        borrows DESC;

                """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                data1 = cursor.fetchall()

            data = []
            if data1 is not None:
                for i in data1:
                    data.append({
                        'grade_name': i[0],
                        'family_name': i[1],
                        'combination_name': i[2],
                        'borrowers': i[3],
                        'students': i[4]
                    })

            serializer = BorrowerByGradeDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
class AllBorrowersDisplayAPIView(APIView):
    # permission_classes = [IsAuthenticated, ]  # You can add authentication if needed

    def get(self, request, *args, **kwargs):
        try:   
            # SQL query for the specified date range
            sql_query = f"""
                SELECT
                    first_name,
                    last_name,
                    phone1,
                    email,
                    grade_name,
                    family_name,
                    combination_name,
                    book_name,
                    isbnumber,
                    category_name,
                    author_name,
                    issuedate,
                    is_student,
                    is_alumni,
                    is_staff,
                    returndate,
                    userprofile_student.id AS student_id,
                    api_user.id AS user_id
                FROM
                    api_user
                INNER JOIN
                    userprofile_issue_book ON userprofile_issue_book.borrower_id = api_user.id
                INNER JOIN
                    userprofile_book ON userprofile_issue_book.book_id = userprofile_book.id
                INNER JOIN
                    userprofile_category ON userprofile_book.category_id = userprofile_category.id
                INNER JOIN
                    userprofile_author ON userprofile_author.id = userprofile_book.author_id
                LEFT JOIN
                    userprofile_student ON api_user.id = userprofile_student.user_id
                LEFT JOIN
                    userprofile_family ON userprofile_student.family_id = userprofile_family.id
                LEFT JOIN
                    userprofile_grade ON userprofile_family.grade_id = userprofile_grade.id
                LEFT JOIN
                    userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id
                ;
            """

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                data1 = cursor.fetchall()

            data = []
            if data1:
                for i in data1:
                    data.append({
                        'first_name': i[0],
                        'last_name': i[1],
                        'phone1': i[2],
                        'email': i[3],
                        'grade_name': i[4],
                        'family_name': i[5],
                        'combination_name': i[6],
                        'book_name': i[7],
                        'isbnumber': i[8],
                        'category_name': i[9],
                        'author_name': i[10],
                        'issuedate': i[11],
                        'is_student':i[12],
                        'is_alumni':i[13],
                        'is_staff':i[14],
                        'returndate': i[15],
                        'student_id': i[16],
                        'user_id': i[17]
                    })

            serializer = AllBorrowersDisplaySerializer(data=data, many=True)
            serializer.is_valid()  # Validate serializer data
            return Response(serializer.data)

        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        
class StudentswithBookReportExportInExcelAPIView(APIView):
    #permission_classes = [IsAuthenticated, ]
    def get_data_from_database(self):
        sql_query = """
           SELECT 
                ROW_NUMBER() OVER (ORDER BY userprofile_grade.grade_name, userprofile_combination.combination_name) AS number,
                last_name,
                first_name,
                studentid,
                grade_name,
                family_name,
                combination_name,
                book_name,
                isbnumber,
                category_name,
                author_name,
                TO_CHAR(
                    CASE 
                        WHEN issuedate LIKE '%T%' THEN TO_TIMESTAMP(issuedate, 'YYYY-MM-DD"T"HH24:MI:SS')
                        WHEN issuedate LIKE '% %' THEN TO_TIMESTAMP(issuedate, 'YYYY-MM-DD HH24:MI:SS')
                        -- Add more formats as needed
                        ELSE NULL
                    END, 
                    'DD Mon, YYYY HH24:MI:SS'
                ) AS issuedate,
                returndate 
            FROM 
                userprofile_grade 
            INNER JOIN 
                userprofile_family ON userprofile_grade.id = userprofile_family.grade_id 
            INNER JOIN 
                userprofile_student ON userprofile_family.id = userprofile_student.family_id 
            INNER JOIN 
                api_user ON userprofile_student.user_id = api_user.id 
            INNER JOIN 
                userprofile_combination ON userprofile_student.combination_id = userprofile_combination.id 
            INNER JOIN 
                userprofile_issue_book ON userprofile_issue_book.borrower_id = api_user.id 
            INNER JOIN 
                userprofile_book ON userprofile_issue_book.book_id = userprofile_book.id 
            INNER JOIN 
                userprofile_category ON userprofile_book.category_id = userprofile_category.id 
            INNER JOIN 
                userprofile_author ON userprofile_book.author_id = userprofile_author.id 
            WHERE 
                returndate = 'Not yet Returned' 
            ORDER BY 
                grade_name ASC, 
                family_name ASC;

        """
        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            data = cursor.fetchall()

        return data

    def generate_excel(self, data):
        # Create a new Workbook
        wb = Workbook()

        # Get the active worksheet
        ws = wb.active
        ws.append(["No", "Last Name","First Name", "Reg.No", "Grade", "Family", "Class","Book Name","ISBN","Category","Author","Isued Date","Returned Date"])
        # Add rows
        isued_book_data_name = 'students_with_book_report'
        ws.title = isued_book_data_name
        #logging.debug(data)
        
        for row_data in data:
            ws.append(row_data)

        # Save the workbook
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=LFHS_Issued_Books_to_students_data.xlsx'
        wb.save(response)
        return response

    def get(self, request, *args, **kwargs):
        try:
            data = self.get_data_from_database()
            if data:
                return self.generate_excel(data)
            else:
                return Response({'error': 'No data found.'}, status=404)
        except Exception as e:
            # Log the exception or return a custom error response
            return Response({'error': str(e)}, status=500)
        

#Alumni update her/his profile  
class AlumniUpdateProfileView(generics.RetrieveUpdateAPIView):
    #permission_classes = [IsAuthenticated, ]
    queryset = Alumni.objects.all()
    serializer_class = AlumniUpdateHisOrHerProfileSerializer

    def get_object(self):
        alumni_id = self.kwargs.get('alumni_id')
        alumn = get_object_or_404(Alumni, id=alumni_id)
        print(alumn)
        return alumn

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)

            return Response(serializer.data)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

#new models (Announcement, Inquiries, frequentlyaskedquestions, Groups )
class AnnouncementViewSet(viewsets.ModelViewSet):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except NotFound as e:
            return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class InquiryViewSet(viewsets.ModelViewSet):
    queryset = Inquiry.objects.all()
    serializer_class = InquirySerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except NotFound as e:
            return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class FrequentlyAskedQuestionViewSet(viewsets.ModelViewSet):
    queryset = FrequentlyAskedQuestion.objects.all()
    serializer_class = FrequentlyAskedQuestionSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except NotFound as e:
            return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except NotFound as e:
            return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

# Data updating process
class UpdateAlumnUploadExcelView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = AlumniUpdatingExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            #sheet_name = 'alumni'  # Hard-code the sheet name here
            
            #sheet_name = 'employement'  # Hard-code the sheet name here
            # Read the specified sheet from the Excel file
            sheet_name = 'studies'  # Hard-code the sheet name here
            sheet_name = 'date_of_birth'  # Hard-code the sheet name here
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
            except ValueError:
                return Response({'error': f'Sheet name "{sheet_name}" does not exist in the uploaded file.'},
                                status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
            
            # Print the first few rows and column names for debugging
            print("Columns in the DataFrame:", df.columns)
            print(df.head())
            
            for index, row in df.iterrows():
                try:
                    alumni = Alumni.objects.get(reg_number=row.get('reg_number'))
                    
                    alumni.did_you_born_in_rwanda = row.get('did_you_born_in_rwanda', alumni.did_you_born_in_rwanda)
                    alumni.place_of_birth_district_or_country = row.get('place_of_birth_district_or_country', alumni.place_of_birth_district_or_country)
                    alumni.place_of_birth_sector_or_city = row.get('place_of_birth_sector_or_city', alumni.place_of_birth_sector_or_city)
                    alumni.date_of_birth = row.get('date_of_birth', alumni.date_of_birth)
                    alumni.save()
                
                except Alumni.DoesNotExist:
                    return Response({f"Alumni record for user with email {row.get('email')} does not exist."})
                    #continue
                except Exception as e:
                    return Response({f"An error occurred while processing row {index} email: {e}"})
                    #continue
                    
                
            
            return Response({'success': 'Database has been updated.'}, status=status.HTTP_200_OK)
        
        # Print serializer errors for debugging
        #print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    
# class UpdateAlumnUploadExcelView(APIView):
#     def post(self, request, *args, **kwargs):
#         serializer = AlumniUpdatingExcelUploadSerializer(data=request.data)
#         if serializer.is_valid():
#             excel_file = serializer.validated_data['file']
#             #sheet_name = 'alumni'  # Hard-code the sheet name here
            
#             #sheet_name = 'employement'  # Hard-code the sheet name here
#             # Read the specified sheet from the Excel file
#             sheet_name = 'studies'  # Hard-code the sheet name here
#             sheet_name = 'date_of_birth'  # Hard-code the sheet name here
#             try:
#                 df = pd.read_excel(excel_file, sheet_name=sheet_name)
#             except ValueError:
#                 return Response({'error': f'Sheet name "{sheet_name}" does not exist in the uploaded file.'},
#                                 status=status.HTTP_400_BAD_REQUEST)
#             except Exception as e:
#                 return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
            
#             # Print the first few rows and column names for debugging
#             print("Columns in the DataFrame:", df.columns)
#             print(df.head())
            
#             for index, row in df.iterrows():
#                 try:
#                     #user = User.objects.get(email=row.get('email'))
#                     alumni = Alumni.objects.get(reg_number=row.get('reg_number'))
#                     """ if not pd.isna(row.get('level')):
#                         # Create new Study record for the alumni
#                         study_data = {
#                             'alumn': alumni,
#                             'level': row.get('level'),
#                             'degree': row.get('degree', ''),
#                             'university': row.get('university', ''),
#                             'country': row.get('country', ''),
#                             'city': row.get('city', ''),
#                             'scholarship': row.get('scholarship'),
#                             'scholarship_details': row.get('scholarship_details', ''),
#                             'status': row.get('status')
#                         }
#                         Studie.objects.create(**study_data)
#                         #print(study_data)
#                     if not pd.isna(row.get('level2')):
#                         study_data2 = {
#                         'alumn': alumni,
#                         'level': row.get('level2'),
#                         'degree': row.get('degree2', ''),
#                         'university': row.get('university2', ''),
#                         'country': row.get('country2', ''),
#                         'city': row.get('city2', ''),
#                         'scholarship': row.get('scholarship2'),
#                         'scholarship_details': row.get('scholarship_details2', ''),
#                         'status': row.get('status2')
#                         }
#                         Studie.objects.create(**study_data2)
#                         #print(study_data2)
#                     if not pd.isna(row.get('level3')):
#                         study_data3 = {
#                         'alumn': alumni,
#                         'level': row.get('level3'),
#                         'degree': row.get('degree3', ''),
#                         'university': row.get('university3', ''),
#                         'country': row.get('country3', ''),
#                         'city': row.get('city3', ''),
#                         'scholarship': row.get('scholarship3'),
#                         'scholarship_details': row.get('scholarship_details3', ''),
#                         'status': row.get('status3')
#                         }
#                         Studie.objects.create(**study_data3)
#                         #print(study_data3) """
                        
#                     alumni.date_of_birth = row.get('date_of_birth', alumni.date_of_birth)
#                     alumni.did_you_born_in_rwanda = row.get('did_you_born_in_rwanda', alumni.did_you_born_in_rwanda)
#                     alumni.place_of_birth_district_or_country = row.get('place_of_birth_district_or_country', alumni.place_of_birth_district_or_country)
#                     alumni.place_of_birth_sector_or_city = row.get('place_of_birth_sector_or_city', alumni.place_of_birth_sector_or_city)
#                     alumni.currresidence_in_rwanda = row.get('currresidence_in_rwanda', alumni.currresidence_in_rwanda)
#                     alumni.currresidence_district_or_country = row.get('currresidence_district_or_country', alumni.currresidence_district_or_country)
#                     alumni.currresidence_sector_or_city = row.get('currresidence_sector_or_city', alumni.currresidence_sector_or_city)
#                     if(row.get('other_email')!=''):
#                         alumni.other_emails =row.get('other_email', alumni.other_emails)
#                     if(row.get('other_phones')!=''):
#                         alumni.other_phones = row.get('other_phones', alumni.other_phones)
#                     alumni.marital_status = row.get('marital_status', alumni.marital_status)
#                     alumni.kids = row.get('kids', alumni.kids) """
                    
#                     alumni.save()
                
#                 except User.DoesNotExist:
#                     return Response({f"User with email {row.get('email')} does not exist."})
#                     #continue
#                 except Alumni.DoesNotExist:
#                     return Response({f"Alumni record for user with email {row.get('email')} does not exist."})
#                     #continue
#                 except Exception as e:
#                     #,{row.get('email')} :1. {row.get('level')},{row.get('degree')},{row.get('university')},{row.get('country')},{row.get('city')},{row.get('scholarship')},{row.get('scholarship_details')}, {row.get('status')}:2. {row.get('email')} {row.get('level2')},{row.get('degree2')},{row.get('university2')},{row.get('country2')},{row.get('city2')},{row.get('scholarship2')},{row.get('scholarship_details2')}, {row.get('status2')},3.{row.get('email')} {row.get('level3')},{row.get('degree3')},{row.get('university3')},{row.get('country3')},{row.get('city3')},{row.get('scholarship3')},{row.get('scholarship_details3')}, {row.get('status3')}
#                     return Response({f"An error occurred while processing row {index} email: {e}"})
#                     #continue
                    
                
            
#             return Response({'success': 'Database has been updated.'}, status=status.HTTP_200_OK)
        
#         # Print serializer errors for debugging
#         #print(serializer.errors)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class EmploymentdraftFilter(filters.FilterSet):
    alumn = filters.NumberFilter(field_name='alumn', lookup_expr='exact')

    class Meta:
        model = Employmentdraft
        fields = ['alumn']

class EmploymentdraftViewSet(viewsets.ModelViewSet):
    queryset = Employmentdraft.objects.all()
    serializer_class = EmploymentdraftSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = EmploymentdraftFilter

    def list(self, request, *args, **kwargs):
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = Employmentdraft1Serializer(instance)
            return Response(serializer.data)
        except Employmentdraft.DoesNotExist:
            raise NotFound(detail="Employmentdraft not found", code=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Employmentdraft.DoesNotExist:
            raise NotFound(detail="Employmentdraft not found", code=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StudiedraftFilter(filters.FilterSet):
    alumn_id = filters.NumberFilter(field_name='alumn_id')

    class Meta:
        model = Studiedraft
        fields = ['alumn_id']

class StudiedraftViewSet(viewsets.ModelViewSet):
    queryset = Studiedraft.objects.all()
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = StudiedraftFilter

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return Studiedraft1Serializer
        return StudiedraftSerializer

    def list(self, request, *args, **kwargs):
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Studiedraft.DoesNotExist:
            raise NotFound(detail="Studiedraft not found", code=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Studiedraft.DoesNotExist:
            raise NotFound(detail="Studiedraft not found", code=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        
#school time table

class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

class TimeSlotsViewSet(viewsets.ModelViewSet):
    queryset = TimeSlots.objects.all()
    serializer_class = TimeSlotsSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

class GradeTimeSlotsViewSet(viewsets.ModelViewSet):
    queryset = GradeTimeSlots.objects.all()
    serializer_class = GradeTimeSlotsSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

class TeacherCombinationGradeSubjectViewSet(viewsets.ModelViewSet):
    queryset = TeacherCombinationGradeSubject.objects.all()
    serializer_class = TeacherCombinationGradeSubjectSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)
        
class AcademicViewSet(viewsets.ModelViewSet):
    queryset = Academic.objects.all()
    serializer_class = AcademicSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, *args, **kwargs):
        try:
            return super().retrieve(request, *args, **kwargs)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer

    # Retrieve all rooms (GET /rooms/)
    def list(self, request, *args, **kwargs):
        try:
            rooms = Room.objects.all()
            serializer = RoomSerializer(rooms, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Retrieve a single room (GET /rooms/<id>/)
    def retrieve(self, request, pk=None, *args, **kwargs):
        try:
            room = Room.objects.get(pk=pk)
            serializer = RoomSerializer(room)
            return Response(serializer.data)
        except Room.DoesNotExist:
            raise NotFound("Room not found")
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Create a new room (POST /rooms/)
    def create(self, request, *args, **kwargs):
        try:
            serializer = RoomSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # Update an existing room (PUT /rooms/<id>/)
    def update(self, request, pk=None, *args, **kwargs):
        try:
            room = Room.objects.get(pk=pk)
            serializer = RoomSerializer(room, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)
        except Room.DoesNotExist:
            raise NotFound("Room not found")
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # Delete a room (DELETE /rooms/<id>/)
    def destroy(self, request, pk=None, *args, **kwargs):
        try:
            room = Room.objects.get(pk=pk)
            room.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Room.DoesNotExist:
            raise NotFound("Room not found")
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#Attendance management System   
class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()

    def get_serializer_class(self):
        # Use AttendanceDetailSerializer for listing and retrieving
        if self.action in ['list', 'retrieve']:
            return AttendanceDetailSerializer
        # Use AttendanceSerializer for creating and updating
        return AttendanceSerializer

    def list(self, request):
        try:
            # Order queryset by 'date' (ascending order)
            attendances = self.get_queryset().order_by('date')  # Use '-date' for descending order
            serializer = self.get_serializer(attendances, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, pk=None):
        try:
            attendance = self.get_object()
            serializer = self.get_serializer(attendance)
            return Response(serializer.data)
        except Attendance.DoesNotExist:
            return Response({'error': 'Attendance not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request):
        try:
            serializer = AttendanceSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, pk=None):
        try:
            attendance = self.get_object()
            serializer = AttendanceSerializer(attendance, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Attendance.DoesNotExist:
            return Response({'error': 'Attendance not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, pk=None):
        try:
            attendance = self.get_object()
            attendance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Attendance.DoesNotExist:
            return Response({'error': 'Attendance not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
       

#English Access Program 
class EapViewSet(viewsets.ModelViewSet):
    queryset = Eap.objects.all()
    serializer_class = EapSerializer

class EapAttendanceViewSet(viewsets.ModelViewSet):
    queryset = EapAttendance.objects.all()
    serializer_class = EapAttendanceSerializer

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        
#new way of taking attendance
logger = logging.getLogger(__name__)

# Define the TimetableFilter for filtering in the viewset
class TimetableFilter(filters.FilterSet):
    academic = filters.NumberFilter(field_name='academic__id', required=True)
    day_of_week = filters.CharFilter(field_name='gradetimeslots__day_of_week', required=True)
    teacher = filters.NumberFilter(field_name='teacher__id', required=True)
    date = filters.DateFilter(method='filter_by_date', required=False)

    class Meta:
        model = TeacherCombinationGradeSubject
        fields = ['academic', 'day_of_week', 'teacher', 'date']

    def filter_by_date(self, queryset, name, value):
        if value:
            attendance_taken = AttendanceTaken.objects.filter(date=value)
            if not attendance_taken.exists():
                return queryset
            return queryset.filter(id__in=attendance_taken.values_list('teachercombinationgradesubject_id', flat=True))
        return queryset

# Define the TimetableViewSet with error handling
class TimetableViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = TimetableSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = TimetableFilter

    def get_queryset(self):
        try:
            academic_id = self.request.query_params.get('academic')
            day_of_week = self.request.query_params.get('day_of_week')
            teacher_id = self.request.query_params.get('teacher')
            date = self.request.query_params.get('date')

            # Ensure required parameters are provided
            if not all([academic_id, day_of_week, teacher_id]):
                return TeacherCombinationGradeSubject.objects.none()

            queryset = TeacherCombinationGradeSubject.objects.select_related(
                'gradetimeslots__grade',
                'gradetimeslots__timeslots',
                'combination',
                'subject',
                'room'
            ).filter(
                academic_id=academic_id,
                gradetimeslots__day_of_week=day_of_week,
                teacher_id=teacher_id
            ).order_by('gradetimeslots__timeslots__start_time')

            if date:
                try:
                    parsed_date = datetime.strptime(date, '%Y-%m-%d').date()
                    attendance_subquery = AttendanceTaken.objects.filter(
                        teachercombinationgradesubject=OuterRef('pk'),
                        date=parsed_date
                    ).values('id')[:1]
                    queryset = queryset.annotate(attendancetaken_id=Subquery(attendance_subquery))
                except ValueError:
                    raise ValueError('Invalid date format. Expected format: YYYY-MM-DD.')

            return queryset

        except Exception as e:
            logger.error(f"Error in get_queryset: {e}")
            raise e

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            date = request.query_params.get('date')
            serializer = self.get_serializer(queryset, many=True, context={'date': date})
            return Response(serializer.data)

        except ValueError as ve:
            return Response(
                {'error': f'Value error: {str(ve)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Unexpected error in list: {e}")
            return Response(
                {'error': 'An internal server error occurred.', 'details': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
class AttendanceTakenViewSet(viewsets.ModelViewSet):
    queryset = AttendanceTaken.objects.all()
    serializer_class = AttendanceTakenSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        attendance_taken = serializer.save()

        # Start with no absentees
        attendance_taken.absentees.clear()
        attendance_taken.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update_absenteeism(self, request, pk=None):
        attendance_taken = self.get_object()
        absenteeism_data = request.data.get('absenteeism', None)

        if absenteeism_data:
            # Create Absenteeism record
            absenteeism_serializer = AbsenteeismSerializer(data=absenteeism_data)
            absenteeism_serializer.is_valid(raise_exception=True)
            absenteeism = absenteeism_serializer.save()

            # Update AttendanceTaken with the new absenteeism
            attendance_taken.absentees.add(absenteeism)
            attendance_taken.save()

            return Response({"message": "AttendanceTaken updated with Absenteeism."}, status=status.HTTP_200_OK)

class AbsenteeismViewSet(viewsets.ModelViewSet):
    queryset = Absenteeism.objects.all()
    serializer_class = AbsenteeismSerializer

class AbsenteeismCommentViewSet(viewsets.ViewSet):
    def create(self, request):
        # Create Absenteeism
        absenteeism_data = {
            "student": request.data.get("student"),
            "status": request.data.get("status"),
        }
        absenteeism_serializer = AbsenteeismSerializer(data=absenteeism_data)
        absenteeism_serializer.is_valid(raise_exception=True)
        absenteeism = absenteeism_serializer.save()

        # Create AttendanceComment if provided
        comment_text = request.data.get('comment')
        if comment_text:
            start_time = request.data.get('start_time')
            end_time = request.data.get('end_time')  # This can be null
            comment_data = {
                "comment": comment_text,
                "start_time": start_time,
                "end_time": end_time,
            }
            comment_serializer = AttendanceCommentSerializer(data=comment_data)
            comment_serializer.is_valid(raise_exception=True)
            comment = comment_serializer.save()

            # Link the comment to the absenteeism record
            absenteeism.school_comments.add(comment)
            absenteeism.save()

        return Response({
            "absenteeism": absenteeism_serializer.data,
            "comment": comment_serializer.data if comment_text else None
        }, status=status.HTTP_201_CREATED)

    def add_comment(self, request, absenteeism_id):
        try:
            absenteeism = Absenteeism.objects.get(id=absenteeism_id)
            comment_text = request.data.get('comment')
            start_time = request.data.get('start_time')
            end_time = request.data.get('end_time')  # This can be null

            comment_data = {
                "comment": comment_text,
                "start_time": start_time,
                "end_time": end_time,
            }
            comment_serializer = AttendanceCommentSerializer(data=comment_data)
            comment_serializer.is_valid(raise_exception=True)
            comment = comment_serializer.save()

            # Link the comment to the absenteeism record
            absenteeism.school_comments.add(comment)
            absenteeism.save()

            return Response({"message": "Comment added successfully.", "comment": comment_serializer.data}, status=status.HTTP_201_CREATED)
        except Absenteeism.DoesNotExist:
            return Response({"error": "Absenteeism record not found."}, status=status.HTTP_404_NOT_FOUND)

class AttendanceCommentViewSet(viewsets.ModelViewSet):
    queryset = AttendanceComment.objects.all()
    serializer_class = AttendanceCommentSerializer

    def update(self, request, pk=None):
        try:
            comment = self.get_object()
            comment_serializer = AttendanceCommentSerializer(comment, data=request.data, partial=True)
            comment_serializer.is_valid(raise_exception=True)
            comment = comment_serializer.save()
            return Response({"message": "Comment updated successfully.", "comment": comment_serializer.data}, status=status.HTTP_200_OK)
        except AttendanceComment.DoesNotExist:
            return Response({"error": "Comment not found."}, status=status.HTTP_404_NOT_FOUND)
        
class StudentListView(generics.ListAPIView):
    serializer_class = StudentListSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['tcgs_id'] = self.kwargs.get('tcgs_id')
        context['date'] = self.request.query_params.get('date')
        return context

    def get_queryset(self):
        tcgs_id = self.kwargs.get('tcgs_id')
        
        # Get the TeacherCombinationGradeSubject instance
        tcgs = get_object_or_404(TeacherCombinationGradeSubject, id=tcgs_id)
        
        # Get students who match both combination_id and grade_id
        students = Student.objects.select_related(
            'user',
            'combination',
            'family__grade'
        ).prefetch_related(
            'absenteeism_set',
            'absenteeism_set__school_comments'
        ).filter(
            combination_id=tcgs.combination_id,
            family__grade_id=tcgs.gradetimeslots.grade_id
        ).order_by('user__last_name', 'user__first_name')

        # Annotate each student with the tcgs_id
        for student in students:
            student.tcgs_id = tcgs_id

        return students

    def list(self, request, *args, **kwargs):
        try:
            tcgs_id = self.kwargs.get('tcgs_id')
            date = request.query_params.get('date')
            
            if date:
                try:
                    # Validate date format
                    datetime.strptime(date, '%Y-%m-%d')
                except ValueError:
                    return Response(
                        {'error': 'Invalid date format. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            tcgs = get_object_or_404(TeacherCombinationGradeSubject, id=tcgs_id)
            
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            # Get attendance information
            attendance_info = None
            if date:
                attendance = AttendanceTaken.objects.filter(
                    teachercombinationgradesubject_id=tcgs_id,
                    date=date
                ).first()
                if attendance:
                    attendance_info = {
                        'id': attendance.id,
                        'date': attendance.date
                    }
            
            response_data = {
                'combination_id': tcgs.combination_id,
                'tcgs_id': tcgs_id,
                'date': date,
                'attendance': attendance_info,
                'students': serializer.data
            }
            
            return Response(response_data)
        except TeacherCombinationGradeSubject.DoesNotExist:
            return Response(
                {'error': 'TeacherCombinationGradeSubject not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )